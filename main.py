from pprint import pprint
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
import os
from start import *


def Legend(table, cursor):
    def getName(el, cursor):
        cursor.execute(
            'SELECT Name_module FROM tbl_module where id_module LIKE %s', (el[0],))
        name = cursor.fetchone()[0]
        return [name, el[1], el[0]]

    legend = []

    for i in range(len(table)):
        # [name, sum_zet, module]
        
        for el in table[i]:
            res = list(filter(lambda x: x[0] == el['module_color'], legend))

            if res != []:
                index = legend.index(res[0])
                legend[index][1] += el['zet']
            else:
                legend.append([el['module_color'], el['zet']])

    if len(legend) <= 3:
        legend = []
        # [name, sum_zet, module]
        for i in range(len(table)):        
            for el in table[i]:
                res = list(filter(lambda x: x[0] == el['block'], legend))

                if res != []:
                    index = legend.index(res[0])
                    legend[index][1] += el['zet']
                else:
                    legend.append([el['block'], el['zet']])
        for i in range(len(legend)):
            legend[i].append(i)

        
        legend.sort(key=lambda x: x[0])
    else:
        legend = list(map(lambda x: getName(x, cursor), legend))
    
    return legend


# Формируем карту excel и сохраняем ее в папку static/temp
def saveMap(aup, cursor, static, **kwargs):

    cur = cursor

    cur.execute("SELECT id_aup, file FROM tbl_aup WHERE num_aup LIKE %s", (aup,))
    id_aup, filename_map = cursor.fetchall()[0]
    
    filename_map_down = f"КД {filename_map}"
    filename_map = os.path.join(static, 'temp', f"КД {filename_map}")

    table, legend = Table(aup, cur, **kwargs)
    ws, wk = CreateMap(filename_map)

    ws.merge_cells(f'A1:{chr(ord("A") + len(table))}1')
    header = Header(aup, cursor)


    header = f'''
        КАРТА ДИСЦИПЛИН УЧЕБНОГО ПЛАНА от {header[4]}
        по направлению подготовки {header[0]}
        Профиль: {header[1]}, {header[2]} год набора, {header[3]}
        АУП: {aup}
    '''
    for i in range(len(table)):
        ws[chr(ord("B")+i)+"2"] = str(i+1) + " семестр"
        ws[chr(ord("B")+i)+"2"].style = 'standart'

    ws['A1'].style = 'standart'
    ws['A1'] = header

    for i in range(len(table)):
        merged = 0
        for el in table[i]:
            column = chr(ord("B") + i)
            cell = f"{column}{3+merged}"

            ws[cell] = el['discipline']
            ws[cell].style = "standart"

            color = el['module_color'].replace('#', '')

            r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            gray = (r + g + b)/3

            if gray < 140:
                ws[cell].font = Font(color="FFFFFF", size=12)
            else:
                ws[cell].font = Font(color="000000", size=12)

            ws[cell].fill = PatternFill(start_color=str(
                color), end_color=str(color), fill_type='solid')

            if el['zet'] < 1:
                el['zet'] = 1.0

            merge_range = f"{cell}:{column}{3+merged + round(el['zet'])-1}"
            ws.merge_cells(merge_range)

            merged += round(el['zet'])

        ws['A40'] = 'З.Е'
        ws['A40'].style = 'standart'
        ws['B40'] = 'МОДУЛИ:'
        ws['B40'].style = 'standart'
        ws.merge_cells('B40:C40')

        # Вывод легенды в КД excel
        sum_zet = 0.0
        for i, el in enumerate(legend):
            cellA = f"A{41+i}"
            cellB = f"B{41+i}"

            sum_zet += el[1]

            ws[cellA].style = ws[cellB].style = 'standart'
            ws[cellA] = el[1]
            ws[cellB] = el[0]
            ws[cellB].alignment = Alignment(
                horizontal='left', vertical='center', wrapText=True)

            color = el[2].replace('#', '')
            r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            gray = (r + g + b)/3

            if gray < 140:
                ws[cellB].font = Font(color="FFFFFF", size=12)
            else:
                ws[cellB].font = Font(color="000000", size=12)

            ws[cellB].fill = PatternFill(start_color=str(
                color), end_color=str(color), fill_type='solid')
            ws.merge_cells(cellB + ':' + cellB.replace('B', 'C'))

    # сумма зет
    cellA = 'A' + str(40+len(legend) + 1)
    ws[cellA].style = 'standart'
    ws[cellA] = f'Итого: {sum_zet}'

    ws['A'+ str(40+len(legend) + 5)] = f'Карта составлена из файла: {filename_map_down}'

    wk.save(filename=filename_map)
    return filename_map


# Возвращает данные для шапка карты
def Header(aup, Cursor):
    # cursor, connection = connect_to_DateBase(FULLNAME_DB)

    cursor = Cursor

    cursor.execute('SELECT id_op FROM tbl_aup WHERE num_aup LIKE %s', (aup,))
    id_op = cursor.fetchall()[0][0]

    cursor.execute(
        'SELECT year_begin, program_code, id_spec, id_form FROM tbl_op WHERE id_op LIKE %s', (id_op,))
    (year_begin, program_code, id_spec, id_form) = cursor.fetchall()[0]

    cursor.execute(
        'SELECT name_okco FROM spr_okco WHERE program_code LIKE %s', (program_code,))
    program = program_code + " " + cursor.fetchall()[0][0]

    cursor.execute(
        'SELECT name_spec FROM spr_specialization WHERE id_spec LIKE %s', (id_spec,))
    spec = cursor.fetchall()[0][0]

    cursor.execute(
        'SELECT form FROM spr_form_education WHERE id_form LIKE %s', (id_form,))
    form = cursor.fetchall()[0][0] + " форма обучения"

    cursor.execute(
        'SELECT file FROM tbl_aup WHERE num_aup LIKE %s', (aup,))
    date_file = cursor.fetchall()[0][0].split(' ')[-4]

    return [program, spec, year_begin, form, date_file]

# раскраска и сортировка данных в таблице
def colorize(table, legend=None, **kwargs):
    COLOR_SET = 0
    EXPO = 0

    colorsets = [
        [
            '#19535F',
            '#0B7A75',
            '#7B2D26',
            '#2AB7CA',
            '#FE4A49',
            '#84596B',
            '#574D68',
            '#A38560',
            '#7EA172',
            '#E7A977',
            '#80475E',
            '#CC5A71',
            '#F0F757',
            '#3B7080',
            '#C97064',
        ], 
        # Случайные цвета, должен быть последним 
        [
            '#%02x%02x%02x' % (randint(0, 255), randint(0, 255), randint(0, 255)) for _ in range(20)
        ]
    ]

    def expo(colorset, value):
        colorset = [[int(x.replace('#', '')[i:i+2], 16)
                     for i in (0, 2, 4)] for x in colorset]

        def f(x):
            r, g, b = x

            r += value
            r = 255 if r > 255 else r

            g += value
            g = 255 if g > 255 else g

            b += value
            b = 255 if b > 255 else b

            return [r, g, b]

        colorset = list(map(lambda x: f(x), colorset))

        for i, el in enumerate(colorset):
            colorset[i] = '#%02x%02x%02x' % (el[0], el[1], el[2])

        return(colorset)

    if 'colorSet' in kwargs.keys():
        if kwargs['colorSet'] < len(colorsets)-1:
            COLOR_SET = kwargs['colorSet']

    if 'expo' in kwargs.keys():
        EXPO = kwargs['expo']

    colorset = expo(colorsets[COLOR_SET], EXPO)



    # находим количество уникальных модулей
    unique_modules = []
    for index, i in enumerate(table):
        for j in i:
            if not j["module_color"] in unique_modules:
                unique_modules.append(j["module_color"])
    print("[DEBUG] uniquemodules = ", unique_modules)
    # преобразовываем список в список списков, где каждый вложенный список состоит из уникального модуля и количества раз
    # когда он встречается в семестрах

    # добавляем цвета легенде
    unique_modules = list(map(lambda x: [x, 0], unique_modules))

    # проверка наличия модулей в карте (бывают карты без заполненного поля модуль)
    
        
    if len(unique_modules) > 3:

        print("[DEBUG] true  = " )
        if legend:
            for i, el in enumerate(legend):
                index = unique_modules.index([el[2], 0])
                legend[i][2] = colorset[index]

        # считаем количетсво семестров для каждого модуля, в которых он встречается
        for i in range(len(table)):
            checked = []
            for j in range(len(table[i])):
                if table[i][j]["module_color"] in checked:
                    continue
                checked.append(table[i][j]["module_color"])

                for k in range(len(unique_modules)):
                    if unique_modules[k][0] == table[i][j]["module_color"]:
                        unique_modules[k][1] += 1

        # преобразовываем список списков в словарь для более простого доступа
        t = {}
        for el in unique_modules:
            t.update({str(el[0]): el[1]})
        unique_modules = t

        # добавляем в таблицу каждому элементу поле count для сортировки
        for i in range(len(table)):
            for j in range(len(table[i])):
                count = unique_modules[str(table[i][j]['module_color'])]
                table[i][j].update({"count": count})

        for i in range(len(table)):
            # сортируем по полю count и module
            table[i].sort(key=lambda x: (
                x["count"], x["module_color"]), reverse=True)

            

            # перемещаем проектную деятельность в конец каждого семестра и раскрашиваем
            j = 0
            swapped = 0
            while j < len(table[i]) - swapped:
                if table[i][j]["module_color"] == 4:
                    table[i][j]["module_color"] = colorset[list(unique_modules.keys()).index(str(table[i][j]["module_color"]))]
                    table[i].append(table[i].pop(j))
                    j -= 1 
                    swapped += 1
                else:
                    table[i][j]["module_color"] = colorset[list(unique_modules.keys()).index(str(table[i][j]["module_color"]))]
                j += 1
            table[i].sort(key=lambda x: x['block'])
            
    else:
        # добавляем цвета легенде
        if legend:
            for i, el in enumerate(legend):
                legend[i][2] = colorset[el[2]]

        

        # раскрашиваем в соответствии с блоком
        for i in range(len(table)):
            # table[i].sort(key=lambda x: x['block'])
            
            j = 0
            swapped = 0
            while j < len(table[i]) - swapped:
                if table[i][j]["module_color"] in [4,21]:
                    
                    el = table[i][j]
                    if "Блок 1" in el['block']:
                        table[i][j]['module_color'] = colorset[0]
                    
                    elif "Блок 2" in el['block']:
                        table[i][j]['module_color'] = colorset[1]

                    else:
                        table[i][j]['module_color'] = colorset[2]

                        
                    table[i].append(table[i].pop(j))
                    j -= 1
                    swapped += 1
                else:
                    el = table[i][j]
                    if "Блок 1" in el['block']:
                        table[i][j]['module_color'] = colorset[0]
                    
                    elif "Блок 2" in el['block']:
                        table[i][j]['module_color'] = colorset[1]

                    else:
                        table[i][j]['module_color'] = colorset[2]

                j += 1    
            
            table[i].sort(key=lambda x: x['block'])        
            
          


        # # если модулей нет
        # disciplines = []
        # discs_count = []
        # # считаем количество дисциплин
        # for i in table:
        #     for j in i:
        #         if not j['discipline'] in disciplines:
        #             disciplines.append(j['discipline'])
        #             discs_count.append(1)
        #         else:
        #             index = disciplines.index(j['discipline'])
        #             discs_count[index] += 1

        # # удаляем те, у которых количество равно 1
        # popped = 0
        # for i, el in enumerate(discs_count):
        #     if el == 1:
        #         disciplines.pop(i-popped)
        #         popped += 1

        # # раскрашиваем в соответсвии с количеством
        # for i in range(len(table)):
        #     for j in range(len(table[i])):
        #         if table[i][j]['discipline'] in disciplines:
        #             index = disciplines.index(table[i][j]['discipline'])
        #             table[i][j]['module_color'] = colorset[index+1]
        #         else:
        #             table[i][j]['module_color'] = colorset[0]

    return table, legend


# возвращает сформированную таблицу с раскрашенными ячейками
def Table(aup, Cursor, **kwargs):
    """
    Make sql-query by aup and buid discipline map \n
    keyword arguments: \n
    colorSet -- number of color set. Default(0)
    expo -- exposition. Bounds: -255 - only black color, only 255 - white color. Default(0)
    """
    sems = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой",
            "Седьмой", "Восьмой", "Девятый", "Десятый", "Одиннадцатый", "Двенадцатый", 'Тринадцатый', 'Четырнадцатый']
    
    # Условия фильтра, если добавлять категорию, то нужно исправить if
    skiplist = {
        "discipline": [
            'Элективные дисциплины по физической культуре и спорту',
            'Элективные курсы по физической культуре и спорту',
            'Элективная физическая культура',
            'Физическая культура',
        ],
        
        "record_type": [
            "Факультативная",
            "Факультативные",
        ]
    }
    
    cursor = Cursor

    cursor.execute('SELECT id_aup FROM tbl_aup WHERE num_aup LIKE %s', (aup,))
    id_aup = cursor.fetchall()

    if id_aup == []:
        # если в бд нет выгрузки
        print("There is no such aup in data base")
        return None
    else:
        id_aup = id_aup[0][0]

    cursor.execute("SELECT id_module, record_type, discipline, period, zet, block FROM workload WHERE id_aup LIKE %s ORDER BY record_type, discipline, period", (id_aup,))
    workload = cursor.fetchall()

    zet = 0.0
    sumzet = 0.0
    # формируем таблицу
    table = []
    for item in workload:
        moduleID, record_type, discipline, period, zet, block = item
        
        # Фильтрация, тут исправлять if 
        if (len(list(filter(lambda x: x in discipline, skiplist['discipline']))) > 0 or 
                len(list(filter(lambda x: x in record_type, skiplist['record_type']))) > 0):
            continue

        sumzet += zet
        period = period.split()[0]

        # словарь с данными ячейчи
        cell = {
            "module_color": moduleID,
            "block": block,
            "discipline": discipline,
            "term": period,
            "zet": zet
        }

        # добаляем в таблицу недостоющее количество семестров для очередной записи в списке workload
        delta = len(table) - sems.index(period) - 1
        if delta < 0:
            for i in range(-delta):
                table.append([])

        # считаем сумму зет для каждой дисциплины, включая экзамены, лаб.занятия, лекции и т.д.
        for el in table[sems.index(period)]:
            if el['discipline'] == discipline:
                el["zet"] += zet
                break
        else:
            # если такой дисциплины нет, то добовляем в таблицу
            table[sems.index(period)].append(cell)

    leg = Legend(table, cursor)
    return colorize(table, legend=leg, **kwargs)  # раскрашиваем и возвращаем


# функция создает карту и задаем все данные кроме предметов в семестрах, на вход требует имя карты
def CreateMap(filename_map):
    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 40, 40)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=False, size=12)
    border = Side(style='medium', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)
    worksheet.row_dimensions[1].height = 50
    worksheet.row_dimensions[2].height = 41
    worksheet["A2"] = "З.Е."
    worksheet['A2'].style = 'standart'
    for col in range(3, 34):
        worksheet["A" + str(col)] = col - 2
        worksheet["A" + str(col)].style = 'standart'
    for col in range(ord('B'), ord('C')):
        worksheet[chr(col) + str(2)] = str(col - 65) + " семестр"
        worksheet[chr(col) + str(2)].style = 'standart'
    return worksheet, workbook


if __name__ == "__main__":
    # main()
    # print(Table('000016957'))
    # Header('000016957')
    # cur, conn = connect_to_DateBase('C:\\Users\\Dezzy\\Documents\\GitHub\\PD_EP_spring2022(1)\\db.accdb')
    # print(select_to_DataBase(cur, 3))
    # saveMap('000016957')
    # file = '29.03.03_2022_Технологии упаковочного производства.xlsx'
    # saveMap(file)
    # get_Table()
    pass
