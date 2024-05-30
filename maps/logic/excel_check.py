from openpyxl import load_workbook

from maps.models import SprStandard, SprVolumeDegreeZET, AupInfo
from maps.logic.tools import get_maximum_rows, skiplist, sems, timeit
from pandas import read_excel, isna
from maps.logic.tools import check_skiplist


# noinspection PyTypeChecker
@timeit
def integrity_check(file):
    """
        Функция для проверки дисциплин учебного плана на целочисленность зет.
        Считает общий объем по дисциплине за семестр, если сумма не целая - записывает ошибку.
        Возвращает список ошибок.
    """

    data = read_excel(file, sheet_name='Лист2', usecols=[0, 4, 5, 6, 8, 9])

    disciplines_amount = {}
    for row in data.values:
        block, record_type, discipline, period, amount, measure = row

        if not check_skiplist(
                zet_or_hours=amount,
                value_discipline=discipline,
                value_block=block,
                value_record_type=record_type):
            continue

        if not isna(amount):
            amount = float(str(amount).replace(',', '.'))
            amount = int(amount * 100) * (54 if measure == "Недели" else 1)

            key = (discipline, period)
            try:
                disciplines_amount[key] += amount
            except:
                disciplines_amount.update({key: amount})

    errors = []
    for key, value in disciplines_amount.items():
        discipline, period = key
        if not (value / 3600).is_integer():
            errors.append(F'{period}: {discipline} {value / 3600}')

    return errors


# noinspection PyTypeChecker
@timeit
def empty_cels_check(file):
    """
        Функция для проверки выгрузки учебного плана на пустые клетки в столбцах A, B, E, F, G, H, J
    """

    data = read_excel(
        file,
        sheet_name='Лист2',
        usecols=[0, 1, 4, 5, 6, 7, 9],
        names=['A', 'B', 'E', 'F', 'G', 'H', 'J'])

    errors = []
    for i in range(len(data)):
        for column in 'ABEFGHJ':
            if data[column][i] is None or isna(data[column][i]):
                errors.append(f"{column}{i}")

    return errors


@timeit
def layout_of_disciplines(file):
    """
        Компоновка элективных дисциплин по семестрам
    """

    wb = load_workbook(file)
    ws = wb['Лист2']
    max_row = get_maximum_rows(sheet_object=ws)
    for num in range(1, max_row + 1):
        if 'Элективные дисциплины' in ws['E' + str(num)].value:
            temp_num = ws['E' + str(num)].value
            temp_value = ws['F' + str(num)].value
            count = 0
            for i in range(num, max_row + 1):
                count += 1
                if ws['E' + str(i)].value == temp_num and ws['F' + str(i)].value != temp_value:
                    for j in range(1, count):
                        ws['F' + str(i - j)] = temp_value + ' / ' + \
                                               ws['F' + str(i + (j - 1))].value
                        ws['F' + str(i + (j - 1))] = 'None'
                        ws['E' + str(i + (j - 1))] = 'None'
    i = 1
    len = max_row + 1
    while i != len:
        if ws['E' + str(i)].value == 'None':
            ws.delete_rows(i)
            len = len - 1
            continue
        i += 1

    wb.save(file)


@timeit
def check_full_zet_in_plan(file):
    """
        Функция для проверки, чтобы общая сумма ЗЕТ соответствовало норме (30 * кол-во семестров)
    """
    wb = load_workbook(file)
    ws = wb['Лист2']
    column_semester = ws['G']
    column_hours = ws['I']
    column_edizm = ws['J']
    column_record_type = ws['E']
    column_discipline = ws['F']
    column_block = ws['A']
    temp_list = []
    for i in range(1, len(column_semester)):
        if column_semester[i].value not in temp_list:
            temp_list.append(column_semester[i].value)

    ws = wb['Лист1']
    program_code = ws['B6'].value
    standard = format_standard(ws['B9'].value)
    id_standard = SprStandard.query.filter_by(
        type_standard=standard).first().id_standard

    select = SprVolumeDegreeZET.query.filter_by(
        program_code=program_code, id_standard=id_standard).first()
    sum_normal = select.zet

    sum_zet = 0
    for i in range(1, len(column_hours)):
        if (column_hours[i].value is not None and (
                len(list(filter(lambda x: x in column_discipline[i].value, skiplist['discipline']))) == 0 and
                len(list(filter(lambda x: x in column_record_type[i].value, skiplist['record_type']))) == 0 and
                len(list(filter(lambda x: x in column_block[i].value, skiplist['record_type']))) == 0)):

            if column_edizm[i].value == 'Недели':
                try:
                    sum_zet += float(column_hours[i].value.replace(',', '.')) * 54
                except:
                    sum_zet += float(column_hours[i].value) * 54
            else:
                try:
                    sum_zet += float(column_hours[i].value.replace(',', '.'))
                except:
                    sum_zet += float(column_hours[i].value)
    sum_zet /= 36

    if abs(round(sum_zet, 2) - sum_zet) < 0.001:
        sum_zet = round(sum_zet, 2)

    wb.save(file)
    return sum_normal, sum_zet


def format_standard(standard):
    if standard == 'ФГОС3++' or standard == 'ФГОС ВО (3++)':
        standard = 'ФГОС ВО 3++'
    return standard


def excel_check(path, aup, options_check):
    return_err_arr = []

    # Проверка на пустые ячейки ###
    errors = empty_cels_check(path)
    if errors:
        return_err_arr.append({
            "message": f'В документе не заполнены ячейки',
            "cells": errors
        })

        return return_err_arr

    # Проверка на целочисленность ЗЕТ у каждой дисциплины ###
    if options_check['checkboxIntegralityModel']:
        errors = integrity_check(path)
        if errors:
            return_err_arr.append({
                'message': f'Ошибка при подсчете ЗЕТ' + '\n'.join(errors)
            })

            return return_err_arr

    # Проверка, чтобы общая сумма ЗЕТ соответствовало норме (30 * кол-во семестров) ###

    if options_check['checkboxSumModel']:
        sum_normal, sum_zet = check_full_zet_in_plan(path)

        if sum_normal != sum_zet:
            return_err_arr.append({
                'message': f'АУП: {aup} В выгрузке общая сумма ЗЕТ не соответствует норме. ' +
                           f'Норма {sum_normal} ЗЕТ. В карте {sum_zet} ЗЕТ.'
            })

            return return_err_arr

    # Проверка на перезапись
    forced_upload = "forced_upload" in options_check and options_check["forced_upload"]
    if not forced_upload and AupInfo.query.filter_by(num_aup=aup).first():
        return_err_arr.append({
            "message": f'Учебный план № {aup} уже существует.',
            "aup": aup
        })

        return return_err_arr

    # Компановка элективных курсов
    layout_of_disciplines(path)

    return return_err_arr
