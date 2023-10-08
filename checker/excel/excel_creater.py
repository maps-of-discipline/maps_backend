import os

from openpyxl import Workbook, worksheet

from .config import *
from .utils import *
from ..data_classes import Detailed, Report, Test


class ExcelCreator:
    DetailedColumnsWidth = Detailed(
        period_id=2,
        discipline=3,
        min=1,
        max=1,
        value=2,
        result=1,
    )

    ColumnTitles = {
        'period_id': 'Период',
        'discipline': 'Дисциплина',
        'min': 'От',
        'max': 'До',
        'value': 'Значение',
        'result': 'Результат',
    }

    period_id_to_title = {}  # A dictionary to map period_id to titles.

    def __init__(self, path: str, hide_title=False, hide_detailed=False):
        """
        Initialize the ExcelCreator object.

        Parameters:
            path (str): The path where the Excel file will be saved.
        """
        self.hide_title = hide_title
        self.hide_detailed = hide_detailed
        self.path = path
        self.workbook = None
        self.sheet: worksheet | None = None
        self.report: Report | None = None

        # If the delete_previous configuration option is set to True, delete all files in the given path.
        if Config.delete_previous[0]:
            for file in os.listdir(self.path):
                if '.xlsx' not in file:
                    continue

                try:
                    os.remove(os.path.join(self.path, file))
                except PermissionError:
                    continue

    def save_report(self, report: Report, folder: str = '') -> str:
        self.workbook = Workbook()
        self.sheet = self.workbook.worksheets[0]
        self.add_named_styles()

        self.report = report
        self.__setup_dimensions()
        self.__make_header()
        self.__fill_report()

        from datetime import datetime
        folder = self.path + folder
        filename = f'{self.report.aup}_{datetime.today().strftime("%H_%M_%S")}_report.xlsx'

        if not os.path.exists(folder):
            os.makedirs(folder)

        self.workbook.save(folder + filename)
        return folder + filename

    def save_reports(self, reports: list[Report], folder: str = ''):
        return [self.save_report(el, folder) for el in reports]

    def __fill_report(self):
        row_index = 10
        for test in self.report.tests:
            if test.result == True and self.hide_title == True:
                continue
            row_index = self.__write_test_header(test, row_index)
            row_index = self.__write_test_details(test, row_index)

    def __process_detailed(self, key, value):
        if value is None:
            return ""
        elif isinstance(value, bool):
            return 'Соответствует' if value else 'Несоответствует'

        elif isinstance(value, list) and key == 'period_id':
            return ' - '.join(self.period_id_to_title[el] for el in value)
        elif key == 'period_id':
            return self.period_id_to_title[value]
        elif isinstance(value, list):
            from math import fsum
            return fsum(value)
        else:
            return value

    def __setup_dimensions(self):
        for i in range(10):
            letter = index_to_column_letter(i)
            self.sheet.column_dimensions[letter].width = Config.base_column_width[0]

        for i in range(1000):
            self.sheet.row_dimensions[i].height = Config.base_row_height[0]

    def add_named_styles(self):
        for attr in Styles.__dict__:
            if "__" not in attr:
                self.workbook.add_named_style(Styles.__getattribute__(Styles, attr))

    def __make_header(self):
        # Merging
        self.sheet.merge_cells('A1:J2')
        self.sheet.merge_cells('A3:B8')
        self.sheet.merge_cells('C3:D3')
        self.sheet.merge_cells('C4:D4')
        self.sheet.merge_cells('C5:D5')
        self.sheet.merge_cells('C6:D6')
        self.sheet.merge_cells('C7:D7')
        self.sheet.merge_cells('C8:D8')
        self.sheet.merge_cells('E3:F3')
        self.sheet.merge_cells('E4:J4')
        self.sheet.merge_cells('E5:J5')
        self.sheet.merge_cells('E6:J6')
        self.sheet.merge_cells('E7:J7')
        self.sheet.merge_cells('E8:G8')
        self.sheet.merge_cells('H3:J3')
        self.sheet.merge_cells('H8:I8')
        self.sheet.merge_cells('A9:C9')

        # Fill data
        self.sheet['A1'] = 'Протокол проверки'
        self.sheet['C3'] = 'Ауп:'
        self.sheet['E3'] = self.report.aup
        self.sheet['C4'] = 'Факультет:'
        self.sheet['E4'] = self.report.faculty
        self.sheet['C5'] = 'Направление:'
        self.sheet['E5'] = self.report.program
        self.sheet['C6'] = 'Профиль:'
        self.sheet['E6'] = self.report.profile
        self.sheet['C7'] = 'Форма обучения:'
        self.sheet['E7'] = self.report.education_form
        self.sheet['C8'] = 'Дата проверки:'
        self.sheet['E8'] = self.report.check_date
        self.sheet['G3'] = 'ОКСО:'
        self.sheet['H3'] = self.report.okso
        self.sheet['H8'] = 'Результат проверки:'
        self.sheet['J8'] = 'Соответствует' if self.report.result else "Несоответствует"
        self.sheet['A9'] = 'Название проверки'
        self.sheet['D9'] = 'От'
        self.sheet['E9'] = 'До'
        self.sheet['F9'] = 'Значение'
        self.sheet['G9'] = 'Ед. Изм'
        self.sheet['J9'] = 'Результат'

        # Styles
        self.sheet['A1'].font = Fonts.h1
        self.sheet['A1'].alignment = Alignments.center
        self.sheet['A1'].fill = Colors.success

        for key in ['C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'G3', 'H8']:
            self.sheet[key].font = Fonts.h2
            self.sheet[key].alignment = Alignments.left

        for key in ['E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'H3']:
            self.sheet[key].font = Fonts.regular
            self.sheet[key].alignment = Alignments.left

        for key in ['A9', 'D9', 'E9', 'F9', 'G9', 'J9', 'H9', 'I9']:
            self.sheet[key].font = Fonts.h2
            self.sheet[key].alignment = Alignments.center
            self.sheet[key].fill = Colors.info

        self.sheet['J8'].font = Fonts.regular
        self.sheet['J8'].alignment = Alignments.center
        self.sheet['J8'].fill = Colors.success if self.report.result else Colors.error

    def __write_test_header(self, test: Test, row_index: int) -> int:
        row_style = Styles.test_success if test.result else Styles.test_error
        self.sheet.merge_cells(f'A{row_index}:C{row_index}')

        self.sheet[f'A{row_index}'] = test.title
        self.sheet[f'A{row_index}'].style = row_style
        self.sheet[f'A{row_index}'].alignment = Alignments.left

        self.sheet[f'D{row_index}'] = test.min
        self.sheet[f'D{row_index}'].style = row_style

        self.sheet[f'E{row_index}'] = test.max
        self.sheet[f'E{row_index}'].style = row_style

        self.sheet[f'F{row_index}'] = test.value
        self.sheet[f'F{row_index}'].style = row_style

        measure = {
            1: 'Часы',
            2: 'Недели',
            3: 'Зет',
            4: 'шт',
            None: ""
        }

        self.sheet[f'G{row_index}'] = measure[test.measure_id]
        self.sheet[f'G{row_index}'].style = row_style

        self.sheet[f'H{row_index}'].style = row_style
        self.sheet[f'I{row_index}'].style = row_style

        self.sheet[f'J{row_index}'] = 'Соответствует' if test.result else 'Несоответствует'
        self.sheet[f'J{row_index}'].style = row_style

        return row_index + 1

    def __write_test_details(self, test: Test, row_index: int) -> int:
        """
        Write detailed information about a test to the Excel sheet.

        Args:
            test (Test): The test object containing detailed information.
            row_index (int): The row index where the details should be written.

        Returns:
            int: The updated row index after writing the details.
        """

        # If test is not detailed, return the original row_index
        if not test.detailed:
            return row_index

        # Initialize variables
        i = 0
        attr_to_column_name = {
            'period_id': 'Период',
            'discipline': 'Дисциплина',
            'min': 'От',
            'max': 'До',
            'value': 'Значение',
            'result': 'Результат',
        }

        # Set up header for the detailed part of the test
        self.sheet.row_dimensions[row_index].height = Config.detailed_row_height[0]
        for attr in self.DetailedColumnsWidth.__dict__:
            merge_from = i
            merge_to = merge_from + self.DetailedColumnsWidth.__getattribute__(attr) - 1
            i += self.DetailedColumnsWidth.__getattribute__(attr)
            merge_range = f'{index_to_column_letter(merge_from)}{row_index}:{index_to_column_letter(merge_to)}{row_index}'
            self.sheet.merge_cells(merge_range)

            # Write column name and apply style
            self.sheet[index_to_column_letter(merge_from) + str(row_index)] = attr_to_column_name[attr]
            self.sheet[index_to_column_letter(merge_from) + str(row_index)].style = Styles.detailed_header

        # Write details of the test
        for row in test.detailed:
            if self.hide_detailed == True and row.result == True:
                continue
            row_index += 1
            self.sheet.row_dimensions[row_index].height = Config.detailed_row_height[0]
            i = 0
            style = Styles.detailed_success if row.result else Styles.detailed_error
            for attr in self.DetailedColumnsWidth.__dict__:
                merge_from = i
                merge_to = merge_from + self.DetailedColumnsWidth.__getattribute__(attr) - 1
                i += self.DetailedColumnsWidth.__getattribute__(attr)
                merge_range = f'{index_to_column_letter(merge_from)}{row_index}:{index_to_column_letter(merge_to)}{row_index}'
                self.sheet.merge_cells(merge_range)

                # Apply style and process attribute values
                self.sheet[index_to_column_letter(merge_from) + str(row_index)].style = style

                if row.__getattribute__(attr) is not None:
                    self.sheet[index_to_column_letter(merge_from) + str(row_index)] = self.__process_detailed(attr,
                                                                                                              row.__getattribute__(
                                                                                                                  attr))

                # Adjust alignment for specific attributes
                if attr == 'discipline' or attr == 'period_id':
                    self.sheet[index_to_column_letter(merge_from) + str(row_index)].alignment = Alignments.left

        # Return the updated row index
        return row_index + 1
