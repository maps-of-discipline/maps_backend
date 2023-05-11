from math import floor
import os
from random import randint
import openpyxl
import xlsxwriter
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)
from tools import get_maximum_rows

from models import (AupInfo, AupData, Groups, db)
from take_from_bd import create_json_print

ROW_START_DISCIPLINES = 5
QUANTITY_HEADER_ROWS = 3


def makeLegend(wb, table):
    ws = wb.create_sheet('Legend')
    ws['A1'].value = 'ЗЕТ'
    ws['A1'].style = 'standart'
    ws['B1'].value = 'Группа'
    ws['B1'].style = 'standart'
    ws.column_dimensions["B"].width = 60.0
    groups = Groups.query.all()
    table_dict = {}
    group_dict = {}
    for group in groups:
        group_dict[group.id_group] = {'name': group.name_group, 'color': group.color}
    for column in table:
        for item in column:
            if item['id_group'] in table_dict:
                table_dict[item['id_group']] += item['zet']
            else:
                table_dict[item['id_group']] = item['zet']
    sum_zet = 0
    for i, key_value in enumerate(table_dict.items()):
        ws['A' + str(i+2)].value = int(key_value[1])
        ws['A' + str(i+2)].style = 'standart'
        ws['B' + str(i+2)].value = group_dict[key_value[0]]['name']
        sum_zet += int(key_value[1])
        color_text_cell(ws, 'B' + str(i+2), group_dict[key_value[0]]['color'].replace('#', ''))
        ws['A' + str(len(table_dict) + 2)].style = 'standart'
        ws['A' + str(len(table_dict) + 2)].value = 'Итого: ' + str(sum_zet)
    # return legend


def saveMap(aup, static, **kwargs):
    aup = AupInfo.query.filter_by(num_aup=aup).first()
    data = AupData.query.filter_by(id_aup=aup.id_aup).all()
    filename_map = aup.file
    filename_map_down = f"КД {filename_map}"
    filename_map = os.path.join(static, 'temp', f"КД {filename_map}")

    table = create_json_print(data)
    max_zet = find_max_zet_excel(table)
    table = add_table_to_arr_and_sort(table['data'])
    ws, wb = CreateMap(filename_map, max_zet, len(table))

    header = Header(aup)
    header1 = f'''КАРТА ДИСЦИПЛИН УЧЕБНОГО ПЛАНА'''
    header2 = f'''Направление подготовки: {header[0]}. Профиль: {header[1]}, {header[2]}. Год набора, {header[3]}. АУП: {aup.num_aup}'''
    ws['A1'].style = 'standart'
    ws['A1'] = header1
    ws['A2'].style = 'standart'
    ws['A2'] = header2

    for row_header in range(1, 3):
        ws.merge_cells(
            f'A{row_header}:{chr(ord("A") + len(table))}{row_header}')

    for width_border in range(1, len(table)+1):
        ws[f"{chr(ord('A') + width_border)}1"].style = 'standart'
        ws[f"{chr(ord('A') + width_border)}2"].style = 'standart'

    for course in range(floor(len(table)/2)):
        ws[chr(ord("B")+course*2)+"3"] = str(course+1) + " курс"
        ws[chr(ord("B")+course*2)+"3"].style = 'standart'
        ws.merge_cells(
            f'{chr(ord("B")+course*2)}3:{chr(ord("B")+course*2+1)}3')
    
    if not (len(table)/2).is_integer():
        ws[chr(ord("B")+floor(len(table)/2)*2)+"3"] = str(floor(len(table)/2)+1) + " курс"
        ws[chr(ord("B")+floor(len(table)/2)*2)+"3"].style = 'standart'

    for semester in range(len(table)):
        ws[chr(ord("B")+semester)+"4"] = str(semester+1)
        ws[chr(ord("B")+semester)+"4"].style = 'standart'

    for i in range(len(table)):
        merged = 0
        for el in table[i]:
            column = chr(ord("B") + i)
            cell = f"{column}{ROW_START_DISCIPLINES+merged}"

            ws[cell] = el['discipline']

            color = el['color'].replace('#', '')
            
            # ws[cell].style = 'standart'

            # r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            # gray = (r + g + b)/3

            # if gray < 140:
            #     ws[cell].font = Font(bold=False, size=18, color="FFFFFF")
            # else:
            #     ws[cell].font = Font(bold=False, size=18, color="000000")

            # ws[cell].fill = PatternFill(start_color=str(
            #     color), end_color=str(color), fill_type='solid')

            color_text_cell(ws, cell, color)

            if el['zet'] < 1:
                el['zet'] = 1.0

            merge_range = f"{cell}:{column}{ROW_START_DISCIPLINES+merged + round(el['zet']*2)-1}"
            ws.merge_cells(merge_range)

            merged += round(el['zet']*2)

    makeLegend(wb, table)

    set_print_properties(table, ws, max_zet)

    wb.save(filename=filename_map)
    return filename_map


def color_text_cell(ws, cell, color):
    ws[cell].style = 'standart'
    r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
    gray = (r + g + b)/3

    if gray < 140:
        ws[cell].font = Font(bold=True, size=14, color="FFFFFF")#Dvorf False 18
    else:
        ws[cell].font = Font(bold=True, size=14, color="000000")#Dvorf False 18

    ws[cell].fill = PatternFill(start_color=str(
        color), end_color=str(color), fill_type='solid')
# def make_many_zet_to_one(table):
#     for item in table:
#         item['zet'] = take_sum_zet_in_discipline(item['type'])
#         del item['type']


def find_max_zet_excel(table):
    max_zet = 0
    terms = {}
    for item in table['data']:
        if item['num_col'] not in terms:
            terms[item['num_col']] = item['zet']
            continue
        terms[item['num_col']] += item['zet']
    for _, value in terms.items():
        if value > max_zet:
            max_zet = value
    return int(max_zet)


def take_sum_zet_in_discipline(item):
    value = 0.0
    for i in item:
        value += i['zet']
    return value


def set_print_properties(table, ws, max_zet):
    # Set properties
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    # ws.page_setup.fitToPage = True
    # ws.row_dimensions[1].height = 100
    ws.page_setup.scale = 60
    max_row = get_maximum_rows(sheet_object=ws)
    ws.print_area = 'A1:' + str(alphabet[len(table)]) + str(max_row*2)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        #left=1/3.81, right=1/3.81, top=1/3.81, bottom=1/3.81, header=1/3.81, footer=1/3.81)
        left=0.25, right=0.25, top=0, bottom=0, header=0, footer=0) #Dvorf

    for height_row in range(ROW_START_DISCIPLINES, max_zet + ROW_START_DISCIPLINES):
        ws.row_dimensions[height_row].height = 17 #Dvorf #35

    ws.column_dimensions['A'].width = 5

    for width_column in range(1, len(table)+1):
        ws.column_dimensions[f'{chr(ord("A")+width_column)}'].width = 40
    ###


# Возвращает данные для шапка карты
def Header(aup):
    year_begin = aup.year_beg
    program = aup.name_op.program_code + ' ' + aup.name_op.okco.name_okco
    form = aup.form.form + " форма обучения"
    spec = aup.name_op.name_spec
    # date_file = aup.file.split(' ')[-4]
    return [program, spec, year_begin, form]


# функция создает карту и задаем все данные кроме предметов в семестрах, на вход требует имя карты
def CreateMap(filename_map, max_zet, table_length):
    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 40, 40)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=True, size=14)
    border = Side(style='thick', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)

    for i in range(1, QUANTITY_HEADER_ROWS):
        worksheet.row_dimensions[i].height = 40

    worksheet["A3"].style = 'standart'
    worksheet["A4"].style = 'standart'

    for col in range(1, max_zet + 1):
        merge_range = f"A{col*2+3}:A{col*2+4}"
        worksheet.merge_cells(merge_range)
        worksheet["A" + str(col*2+3)] = col
        worksheet["A" + str(col*2+3)].style = 'standart'
        worksheet["A" + str(col*2+4)].style = 'standart'
    for col in range(ord('B'), ord('C')):
        worksheet[chr(col) + str(5)].style = 'standart'
    return worksheet, workbook


def add_table_to_arr_and_sort(table):
    print(table)
    count_columns = 0
    for item in table:
        if item['num_col'] > count_columns:
            count_columns = item['num_col']

    new_table = [[] for _ in range(count_columns)]

    for item in table:
        new_table[item['num_col'] - 1].append(item)

    for a in range(len(new_table)):
        for i in range(len(new_table[a])-1):
            for j in range(len(new_table[a])-i-1):
                if new_table[a][j]['num_row'] > new_table[a][j+1]['num_row']:
                    new_table[a][j], new_table[a][j +
                                                  1] = new_table[a][j+1], new_table[a][j]
    return (new_table)

