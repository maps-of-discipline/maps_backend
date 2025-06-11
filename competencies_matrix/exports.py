# filepath: competencies_matrix/exports.py
import io
import logging
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DEFAULT_FONT_NAME = 'Times New Roman'
DEFAULT_FONT_SIZE = 11.5 # Лучше использовать целое число, т.к. в openpyxl и LibreOffice число может округляться
TABLE_TITLE_FONT_SIZE = 14

def generate_tf_list_excel_export(selected_data: Dict[str, Any], opop_data: Dict[str, str]) -> bytes:
    """
    Генерирует Excel-файл с перечнем выбранных ОТФ и ТФ в формате,
    максимально похожем на предоставленное изображение (Таблица 1 из ОПОП).

    Args:
        selected_data: Словарь, содержащий данные о выбранных ТФ.
        opop_data: Словарь с данными об образовательной программе для заголовка.
                   Пример: {'direction_code': '09.03.01', 'direction_name': 'Информатика и вычислительная техника', 'profile_name': 'Веб-технологии'}

    Returns:
        bytes: Содержимое Excel-файла в виде байтов.
    """
    logger.info("Starting Excel export generation for TF List (OPOP Table 1 format).")
    wb = Workbook()
    ws = wb.active
    ws.title = "Перечень ТФ"

    # --- Стили ---
    title_font = Font(name=DEFAULT_FONT_NAME, size=TABLE_TITLE_FONT_SIZE, bold=True)
    header_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    cell_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    
    cell_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    
    # Выравнивание: перенос текста, вертикально по центру/верху, горизонтально по центру/левому краю
    alignment_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    alignment_wrap_left_top = Alignment(wrap_text=True, vertical='top', horizontal='left')

    # --- Заголовок над таблицей ---
    title_text = (
        "Таблица 1 – Перечень обобщённых трудовых функций и трудовых функций, соответствующих профессиональной "
        "деятельности выпускника программы бакалавриата по направлению подготовки "
        f"{opop_data.get('direction_code', '[Код]')} {opop_data.get('direction_name', '[Направление]')}, "
        f"профиль «{opop_data.get('profile_name', '[Профиль]')}»"
    )
    ws.append([title_text])
    title_cell = ws['A1']
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center') # Для заголовка выравнивание по центру
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 60 # Примерная высота для заголовка (можно настроить)

    # Пустая строка после заголовка
    ws.append([])
    
    # --- Заголовки таблицы ---
    current_row_idx = 3 # Начинаем с третьей строки
    
    # Добавляем основные заголовки в ячейки
    ws.cell(row=current_row_idx, column=1, value="Код и наименование\nпрофессионального\nстандарта")
    ws.cell(row=current_row_idx, column=2, value="Обобщенные трудовые функции")
    ws.cell(row=current_row_idx, column=5, value="Трудовые функции")

    # Объединение ячеек для основных заголовков
    ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx + 1, end_column=1)
    ws.merge_cells(start_row=current_row_idx, start_column=2, end_row=current_row_idx, end_column=4)
    ws.merge_cells(start_row=current_row_idx, start_column=5, end_row=current_row_idx, end_column=7)

    # Добавление подзаголовков (row=current_row_idx + 1)
    ws.cell(row=current_row_idx + 1, column=2, value="код")
    ws.cell(row=current_row_idx + 1, column=3, value="наименование")
    ws.cell(row=current_row_idx + 1, column=4, value="уровень\nквалификации")
    ws.cell(row=current_row_idx + 1, column=5, value="наименование")
    ws.cell(row=current_row_idx + 1, column=6, value="код")
    ws.cell(row=current_row_idx + 1, column=7, value="уровень\n(подуровень)\nквалификации")
    
    # Применение стилей к заголовкам
    for row in ws.iter_rows(min_row=current_row_idx, max_row=current_row_idx + 1, min_col=1, max_col=7):
        for cell in row:
            if cell.value: # Применяем только к ячейкам с текстом (мерджед могут быть пустыми)
                cell.font = header_font
                cell.border = cell_border
                cell.alignment = alignment_center # Все заголовки по центру

    # Установка ШИРИНЫ КОЛОНОК - ЭТО КРИТИЧЕСКИ ВАЖНО!
    # Увеличиваем значения, чтобы дать место для переноса текста
    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 60
    ws.column_dimensions[get_column_letter(4)].width = 15
    ws.column_dimensions[get_column_letter(5)].width = 60
    ws.column_dimensions[get_column_letter(6)].width = 15
    ws.column_dimensions[get_column_letter(7)].width = 18

    # --- Заполнение данными ---
    data_start_row = current_row_idx + 2
    
    for ps_data in selected_data.get('profStandards', []):
        ps_start_row_in_data = data_start_row # Начальная строка для текущего ПС
        
        for otf_data in ps_data.get('generalized_labor_functions', []):
            otf_start_row_in_data = data_start_row # Начальная строка для текущей ОТФ
            
            for i, tf_data in enumerate(otf_data.get('labor_functions', [])):
                row_values = [
                    "",
                    "",
                    "",
                    "",
                    tf_data.get('name', ''),
                    tf_data.get('code', ''),
                    tf_data.get('qualification_level', '')
                ]
                
                ws.append(row_values)
                
                # Применяем стили и выравнивание к только что добавленной строке
                current_row = ws[ws.max_row]
                for cell in current_row:
                    cell.font = cell_font
                    cell.border = cell_border
                    # Устанавливаем выравнивание
                    if cell.column in [3, 5]: # Колонки "Наименование" ОТФ и ТФ
                        cell.alignment = alignment_wrap_left_top
                    else:
                        cell.alignment = alignment_center
                
                # Устанавливаем автоподбор высоты строки. None означает "авто"
                ws.row_dimensions[ws.max_row].height = None 
                
                data_start_row += 1

            # Слияние ячеек для ОТФ (если в ней было больше одной ТФ)
            if data_start_row - 1 > otf_start_row_in_data:
                # Вставляем значения ОТФ в первую ячейку объединенного диапазона
                ws.cell(row=otf_start_row_in_data, column=2, value=otf_data.get('code', ''))
                ws.cell(row=otf_start_row_in_data, column=3, value=otf_data.get('name', ''))
                ws.cell(row=otf_start_row_in_data, column=4, value=otf_data.get('qualification_level', ''))

                ws.merge_cells(start_row=otf_start_row_in_data, start_column=2, end_row=data_start_row - 1, end_column=2)
                ws.merge_cells(start_row=otf_start_row_in_data, start_column=3, end_row=data_start_row - 1, end_column=3)
                ws.merge_cells(start_row=otf_start_row_in_data, start_column=4, end_row=data_start_row - 1, end_column=4)

        # Слияние ячеек для ПС (если в нем было больше одной строки данных)
        if data_start_row - 1 > ps_start_row_in_data:
            # Вставляем значение ПС в первую ячейку объединенного диапазона
            ws.cell(row=ps_start_row_in_data, column=1, value=f"{ps_data.get('code', '')}\n{ps_data.get('name', '')}")
            ws.merge_cells(start_row=ps_start_row_in_data, start_column=1, end_row=data_start_row - 1, end_column=1)

    # Сохранение в байты
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    logger.info("Excel export for TF List (OPOP Table 1 format) generation finished.")
    return bio.getvalue()