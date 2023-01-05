from math import ceil
import os
from random import randint
import openpyxl
import xlsxwriter
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)
from tools import get_maximum_rows

from models import (AUP, OP, Module, NameOP, SprFaculty, SprFormEducation,
                    SprOKCO, Workload, WorkMap, db)

# Условия фильтра, если добавлять категорию, то нужно исправить if
skiplist = {
    "discipline": [
        'Элективные дисциплины по физической культуре и спорту',
        'Элективные курсы по физической культуре и спорту', 
        'Элективная физическая культура',
        'Общая физическая подготовка',
        'Игровые виды спорта',
        'Неолимпийские виды спорта'
    ],

    "record_type": [
        "Факультативная",
        "Факультативные"
    ]
}

def takeTableForExcel(aup):
    q = WorkMap.query.filter_by(id_aup=aup).all()
    d = dict()
    l = list()
    for i in q:
        a = dict()
        a["id"] = i.id
        a["discipline"] = i.discipline
        a["zet"] = i.zet
        a["id_group"] = i.id_group
        a["num_col"] = i.num_col
        a["num_row"] = i.num_row
        a["disc_color"] = i.disc_color
        l.append(a)
    d["data"] = l
    return d


def Legend(table):
    def getName(el):

        name = Module.query.filter_by(id_module=el[0]).first().name_module
        return [name, el[1], el[0]]

    legend = []

    for i in range(len(table)):
        # [name, sum_zet, module]

        for el in table[i]:
            res = list(filter(lambda x: x[0] == el['module_color'], legend))

            if res != []:
                index = legend.index(res[0])
                legend[index][1] += el['zet']
            else:
                legend.append([el['module_color'], el['zet']])

    if len(legend) <= 3:
        legend = []
        # [name, sum_zet, module]
        for i in range(len(table)):
            for el in table[i]:
                res = list(filter(lambda x: x[0] == el['block'], legend))

                if res != []:
                    index = legend.index(res[0])
                    legend[index][1] += el['zet']
                else:
                    legend.append([el['block'], el['zet']])
        for i in range(len(legend)):
            legend[i].append(i)

        legend.sort(key=lambda x: x[0])
    else:
        legend = list(map(lambda x: getName(x), legend))

    return legend


def saveMap(aup, static, **kwargs):
    select_aup = AUP.query.filter_by(num_aup=aup).first()
    id_aup = select_aup.id_aup
    filename_map = select_aup.file

    filename_map_down = f"КД {filename_map}"
    filename_map = os.path.join(static, 'temp', f"КД {filename_map}")

    _, legend, _ = Table(aup, **kwargs)
    table = takeTableForExcel(aup)
    max_zet = find_max_zet_excel(table)
    table = add_table_to_arr_and_sort(table['data'])
    ws, wk = CreateMap(filename_map, max_zet, len(table))
    

    for row_header in range(1, 3):
        ws.merge_cells(f'A{row_header}:{chr(ord("A") + len(table))}{row_header}')

    for width_border in range(1, len(table)+1):
        ws[f"{chr(ord('A') + width_border)}1"].style = 'standart'
        ws[f"{chr(ord('A') + width_border)}2"].style = 'standart'

    header = Header(aup)
    header1 = f'''КАРТА ДИСЦИПЛИН УЧЕБНОГО ПЛАНА от {header[4]}'''
    header2 = f'''Направление подготовки: {header[0]}. Профиль: {header[1]}, {header[2]}. Год набора, {header[3]}. АУП: {aup}'''

    for course in range(ceil(len(table)/2)):
        ws[chr(ord("B")+course*2)+"3"] = str(course+1) + " курс"
        ws[chr(ord("B")+course*2)+"3"].style = 'standart'
        ws.merge_cells(f'{chr(ord("B")+course*2)}3:{chr(ord("B")+course*2+1)}3')

    for semester in range(len(table)):
        ws[chr(ord("B")+semester)+"4"] = str(semester+1)
        ws[chr(ord("B")+semester)+"4"].style = 'standart'

    ws['A1'].style = 'standart'
    ws['A1'] = header1

    ws['A2'].style = 'standart'
    ws['A2'] = header2

    ROW_START_DISCIPLINES = 5

    for i in range(len(table)):
        merged = 0
        for el in table[i]:
            column = chr(ord("B") + i)
            cell = f"{column}{ROW_START_DISCIPLINES+merged}"

            ws[cell] = el['discipline']
            ws[cell].style = 'standart'

            color = el['disc_color'].replace('#', '')

            r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            gray = (r + g + b)/3

            if gray < 140:
                ws[cell].font = Font(bold=False, size=18, color="FFFFFF")
            else:
                ws[cell].font = Font(bold=False, size=18, color="000000")

            ws[cell].fill = PatternFill(start_color=str(
                color), end_color=str(color), fill_type='solid')

            if el['zet'] < 1:
                el['zet'] = 1.0

            merge_range = f"{cell}:{column}{ROW_START_DISCIPLINES+merged + round(el['zet'])-1}"
            ws.merge_cells(merge_range)

            merged += round(el['zet'])

    legend_on_2nd_sheet(wk, legend, filename_map_down)

    ### Set properties
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    # ws.page_setup.fitToPage = True
    # ws.row_dimensions[1].height = 100
    ws.page_setup.scale = 65
    max_row = get_maximum_rows(sheet_object=ws)
    ws.print_area = 'A1:' + str(alphabet[len(table)]) + str(max_row)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(left=1/3.81, right=1/3.81, top=1/3.81, bottom=1/3.81, header=1/3.81, footer=1/3.81)
    
    for height_row in range(ROW_START_DISCIPLINES, max_zet + ROW_START_DISCIPLINES):
        ws.row_dimensions[height_row].height = 35

    ws.column_dimensions['A'].width = 5

    for width_column in range(1, len(table)+1):
        ws.column_dimensions[f'{chr(ord("A")+width_column)}'].width = 40
    ###

    wk.save(filename=filename_map)
    return filename_map

def add_table_to_arr_and_sort(table):
    print(table)
    new_table = []
    count_columns = 0
    for item in table:
        if item['num_col'] > count_columns:
            count_columns = item['num_col']
    
    for i in range(count_columns + 1):
        new_table.append([])

    for item in table:
        new_table[item['num_col']].append(item)

    for a in range(len(new_table)):
        for i in range(len(new_table[a])-1):
            for j in range(len(new_table[a])-i-1):
                if new_table[a][j]['num_row'] > new_table[a][j+1]['num_row']:
                    new_table[a][j], new_table[a][j+1] = new_table[a][j+1], new_table[a][j]
    return(new_table)

def legend_on_2nd_sheet(wb, legend, filename_map_down):
    # Вывод легенды в КД excel
    ws = wb.create_sheet('Legend')
    ws['A1'] = 'З.Е'
    ws['A1'].style = 'standart'
    ws['B1'] = 'МОДУЛИ:'
    ws['B1'].style = 'standart'
    ws.merge_cells('B1:C1')

    sum_zet = 0.0
    for i, el in enumerate(legend):
        cellA = f"A{2+i}"
        cellB = f"B{2+i}"

        sum_zet += el[1]

        ws[cellA].style = ws[cellB].style = 'standart'
        ws[cellA] = el[1]
        ws[cellB] = el[0]
        ws[cellB].alignment = Alignment(
            horizontal='left', vertical='center', wrapText=True)

        color = el[2].replace('#', '')
        r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        gray = (r + g + b)/3

        if gray < 140:
            ws[cellB].font = Font(color="FFFFFF", size=18)
        else:
            ws[cellB].font = Font(color="000000", size=18)

        ws[cellB].fill = PatternFill(start_color=str(
            color), end_color=str(color), fill_type='solid')
        ws.merge_cells(cellB + ':' + cellB.replace('B', 'C'))

    # сумма зет
    cellA = 'A' + str(1+len(legend) + 1)
    ws[cellA].style = 'standart'
    ws[cellA] = f'Итого: {sum_zet}'

    ws['A' + str(1+len(legend) + 5)
    ] = f'Карта составлена из файла: {filename_map_down}'
    ws.column_dimensions['C'].width = 75

# Возвращает данные для шапка карты
def Header(aup):

    id_op = AUP.query.filter_by(num_aup=aup).first().id_op

    select_op = OP.query.filter_by(id_op=id_op).first()
    year_begin = select_op.duration.year_beg
    program_code = select_op.duration.name_op.program_code
    id_spec = select_op.duration.id_spec
    id_form = select_op.duration.id_form

    program = program_code + " " + \
        SprOKCO.query.filter_by(program_code=program_code).first().name_okco

    spec = NameOP.query.filter_by(id_spec=id_spec).first().name_spec

    form = SprFormEducation.query.filter_by(
        id_form=id_form).first().form + " форма обучения"

    date_file = AUP.query.filter_by(num_aup=aup).first().file.split(' ')[-4]

    return [program, spec, year_begin, form, date_file]



# раскраска и сортировка данных в таблице
def colorize(table, legend=None, **kwargs):
    COLOR_SET = 0
    EXPO = 0

    colorsets = [
        [
            '#5f60ec', #19535F Прошлые цвета (Дворф)
            '#5f99ec', #0B7A75 Прошлые цвета (Дворф)
            '#ec815f', #7B2D26 Прошлые цвета (Дворф)
            '#5fecd2', #2AB7CA Прошлые цвета (Дворф)
            '#ec5f6b', #FE4A49 Прошлые цвета (Дворф)
            '#ef9f90', #84596B Прошлые цвета (Дворф)
            '#8f5fec', #574D68 Прошлые цвета (Дворф)
            '#eca75f', #A38560 Прошлые цвета (Дворф)
            '#5fec8e', #7EA172 Прошлые цвета (Дворф)
            '#d7ec5f', #E7A977 Прошлые цвета (Дворф)
            '#cf5fec', #80475E Прошлые цвета (Дворф)
            '#ec5fd0', #CC5A71 Прошлые цвета (Дворф)
            '#eceb5f', #FFCC00 Прошлые цвета (Дворф)
            '#5fd2ec', #3B7080 Прошлые цвета (Дворф)
            '#ecc95f', #C97064 Прошлые цвета (Дворф)
            '#5fec8e', #008B8B Прошлые цвета (Дворф)
            '#ecc95f', #B8860B Прошлые цвета (Дворф)
            '#8fec5f'  #01B235 Прошлые цвета (Дворф)
        ],
        # Случайные цвета, должен быть последним
        [
            '#%02x%02x%02x' % (randint(0, 255), randint(0, 255), randint(0, 255)) for _ in range(20)
        ]
    ]

    def expo(colorset, value):
        colorset = [[int(x.replace('#', '')[i:i+2], 16)
                     for i in (0, 2, 4)] for x in colorset]

        def f(x):
            r, g, b = x

            r += value
            r = 255 if r > 255 else r

            g += value
            g = 255 if g > 255 else g

            b += value
            b = 255 if b > 255 else b

            return [r, g, b]

        colorset = list(map(lambda x: f(x), colorset))

        for i, el in enumerate(colorset):
            colorset[i] = '#%02x%02x%02x' % (el[0], el[1], el[2])

        return (colorset)

    if 'colorSet' in kwargs.keys():
        if kwargs['colorSet'] < len(colorsets)-1:
            COLOR_SET = kwargs['colorSet']

    if 'expo' in kwargs.keys():
        EXPO = kwargs['expo']

    colorset = expo(colorsets[COLOR_SET], EXPO)

    # находим количество уникальных модулей
    unique_modules = []
    for index, i in enumerate(table):
        for j in i:
            if not j["module_color"] in unique_modules:
                unique_modules.append(j["module_color"])
    print("[DEBUG] uniquemodules = ", unique_modules)
    # преобразовываем список в список списков, где каждый вложенный список состоит из уникального модуля и количества раз
    # когда он встречается в семестрах

    # добавляем цвета легенде
    unique_modules = list(map(lambda x: [x, 0], unique_modules))

    # проверка наличия модулей в карте (бывают карты без заполненного поля модуль)

    if len(unique_modules) > 3:

        print("[DEBUG] true  = ")
        if legend:
            for i, el in enumerate(legend):
                index = unique_modules.index([el[2], 0])
                legend[i][2] = colorset[index]

        # считаем количетсво семестров для каждого модуля, в которых он встречается
        for i in range(len(table)):
            checked = []
            for j in range(len(table[i])):
                if table[i][j]["module_color"] in checked:
                    continue
                checked.append(table[i][j]["module_color"])

                for k in range(len(unique_modules)):
                    if unique_modules[k][0] == table[i][j]["module_color"]:
                        unique_modules[k][1] += 1

        # преобразовываем список списков в словарь для более простого доступа
        t = {}
        for el in unique_modules:
            t.update({str(el[0]): el[1]})
        unique_modules = t

        # добавляем в таблицу каждому элементу поле count для сортировки
        for i in range(len(table)):
            for j in range(len(table[i])):
                count = unique_modules[str(table[i][j]['module_color'])]
                table[i][j].update({"count": count})

        for i in range(len(table)):
            # сортируем по полю count и module
            table[i].sort(key=lambda x: (
                x["count"], x["module_color"]), reverse=True)

            # перемещаем проектную деятельность в конец каждого семестра и раскрашиваем
            j = 0
            swapped = 0
            while j < len(table[i]) - swapped:
                if table[i][j]["module_color"] == 4:
                    table[i][j]["module_color"] = colorset[list(
                        unique_modules.keys()).index(str(table[i][j]["module_color"]))]
                    table[i].append(table[i].pop(j))
                    j -= 1
                    swapped += 1
                else:
                    table[i][j]["module_color"] = colorset[list(
                        unique_modules.keys()).index(str(table[i][j]["module_color"]))]
                j += 1
            table[i].sort(key=lambda x: x['block'])

    else:
        # добавляем цвета легенде
        if legend:
            for i, el in enumerate(legend):
                legend[i][2] = colorset[el[2]]

        # раскрашиваем в соответствии с блоком
        for i in range(len(table)):
            # table[i].sort(key=lambda x: x['block'])

            j = 0
            swapped = 0
            while j < len(table[i]) - swapped:
                if table[i][j]["module_color"] in [4, 21]:

                    el = table[i][j]
                    if "Блок 1" in el['block']:
                        table[i][j]['module_color'] = colorset[0]

                    elif "Блок 2" in el['block']:
                        table[i][j]['module_color'] = colorset[1]

                    else:
                        table[i][j]['module_color'] = colorset[2]

                    table[i].append(table[i].pop(j))
                    j -= 1
                    swapped += 1
                else:
                    el = table[i][j]
                    if "Блок 1" in el['block']:
                        table[i][j]['module_color'] = colorset[0]

                    elif "Блок 2" in el['block']:
                        table[i][j]['module_color'] = colorset[1]

                    else:
                        table[i][j]['module_color'] = colorset[2]

                j += 1

            table[i].sort(key=lambda x: x['block'])

        # # если модулей нет
        # disciplines = []
        # discs_count = []
        # # считаем количество дисциплин
        # for i in table:
        #     for j in i:
        #         if not j['discipline'] in disciplines:
        #             disciplines.append(j['discipline'])
        #             discs_count.append(1)
        #         else:
        #             index = disciplines.index(j['discipline'])
        #             discs_count[index] += 1

        # # удаляем те, у которых количество равно 1
        # popped = 0
        # for i, el in enumerate(discs_count):
        #     if el == 1:
        #         disciplines.pop(i-popped)
        #         popped += 1

        # # раскрашиваем в соответсвии с количеством
        # for i in range(len(table)):
        #     for j in range(len(table[i])):
        #         if table[i][j]['discipline'] in disciplines:
        #             index = disciplines.index(table[i][j]['discipline'])
        #             table[i][j]['module_color'] = colorset[index+1]
        #         else:
        #             table[i][j]['module_color'] = colorset[0]

    return table, legend


sems = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой",
            "Седьмой", "Восьмой", "Девятый", "Десятый", "Одиннадцатый", "Двенадцатый", 'Тринадцатый', 'Четырнадцатый']

# возвращает сформированную таблицу с раскрашенными ячейками
def Table(aup, **kwargs):
    """
    Make sql-query by aup and buid discipline map \n
    keyword arguments: \n
    colorSet -- number of color set. Default(0)
    expo -- exposition. Bounds: -255 - only black color, only 255 - white color. Default(0)
    """
    # sems = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой",
    #         "Седьмой", "Восьмой", "Девятый", "Десятый", "Одиннадцатый", "Двенадцатый", 'Тринадцатый', 'Четырнадцатый']


    id_aup = AUP.query.filter_by(num_aup=aup).first().id_aup

    if id_aup == None:
        # если в бд нет выгрузки
        print("There is no such aup in data base")
        return None

    workload = Workload.query.filter_by(id_aup=id_aup).order_by(
        Workload.record_type.desc(), Workload.discipline.desc(), Workload.period.desc()).all()
    print('--------------------', workload[0].load)
    sumzet = 0.0
    # формируем таблицу
    table = []
    for item in workload:
        # moduleID, record_type, discipline, period, zet, block = item
        moduleID = item.id_module
        record_type = item.record_type
        discipline = item.discipline
        period = item.period
        zet = item.zet
        block = item.block

        # Фильтрация, тут исправлять if
        if (len(list(filter(lambda x: x in discipline, skiplist['discipline']))) > 0 or
                len(list(filter(lambda x: x in record_type, skiplist['record_type']))) > 0 or
                len(list(filter(lambda x: x in block, skiplist['record_type']))) > 0):
            continue

        sumzet += zet
        period = period.split()[0]

        # словарь с данными ячейчи
        cell = {
            "module_color": moduleID,
            "block": block,
            "discipline": discipline,
            "term": period,
            "zet": zet
        }

        # добаляем в таблицу недостоющее количество семестров для очередной записи в списке workload
        delta = len(table) - sems.index(period) - 1
        if delta < 0:
            for i in range(-delta):
                table.append([])

        # считаем сумму зет для каждой дисциплины, включая экзамены, лаб.занятия, лекции и т.д.
        for el in table[sems.index(period)]:
            if el['discipline'] == discipline:
                el['zet'] += zet
                break
        else:
            # если такой дисциплины нет, то добавляем в таблицу
            table[sems.index(period)].append(cell)

    leg = Legend(table)
    print('-------------------', sumzet)
    max_zet = find_max_zet(table)
    print('!!!!!---------!!!!!!!', table)
    table, legend = colorize(table, legend=leg, **kwargs) # раскрашиваем и возвращаем
    print('!!!!!---------!!!!!!!', table)
    return table, legend, max_zet

def find_max_zet(table):
    max_zet = 0
    for column in table:
        temp = 0
        for it in column:
            temp += it['zet']
        if temp > max_zet:
            max_zet = temp
    return int(max_zet)


def find_max_zet_excel(table):
    # print('!=!=!=!=!=!=!')
    # print(table)
    max_zet = 0
    terms = {}
    for item in table['data']:
        if item['num_col'] not in terms:
            terms[item['num_col']] = item['zet']
            continue
        terms[item['num_col']] += item['zet']
    for _, value in terms.items():
        if value > max_zet:
            max_zet = value
    return int(max_zet)

# функция создает карту и задаем все данные кроме предметов в семестрах, на вход требует имя карты
def CreateMap(filename_map, max_zet, table_length):
    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 40, 40)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=False, size=18)
    border = Side(style='thick', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)
    
    QUANTITY_HEADER_ROWS = 3
    for i in range(1, QUANTITY_HEADER_ROWS):
        worksheet.row_dimensions[i].height = 30

    worksheet["A3"].style = 'standart'
    worksheet["A4"].style = 'standart'

    for col in range(5, max_zet + 5):
        worksheet["A" + str(col)] = col - 4
        worksheet["A" + str(col)].style = 'standart'
    for col in range(ord('B'), ord('C')):
        worksheet[chr(col) + str(5)].style = 'standart'
    return worksheet, workbook


# получить все сущестующие карты
def GetMaps(id, name=None):
    res = dict()
    q = db.session.query(
       AUP.num_aup,
       AUP.file, 
    )

    if name != None:
        search = "%{}%".format(name)
        join_query = q.join(OP).join(SprFaculty).filter(AUP.file.like(search))
    else:
        join_query = q.join(OP).join(SprFaculty)

    maps = join_query.filter(SprFaculty.id_faculty == id).all()

    db.session.close()
    db.session.expunge_all()

    # print(maps)



    # for i in range(len(maps)):
    #     print(maps[i].id_aup)
    #     res[maps[i].id_aup] = [maps[i].id_op, maps[i].file, maps[i].num_aup, maps[i].op]

    return maps

# получить список всех факултетов
def GetAllFaculties():
    fac = SprFaculty.query.all()
    return fac

if __name__ == "__main__":

    pass
