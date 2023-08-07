from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill


class Fonts:
    h1 = Font(size=20, bold=True)
    h2 = Font(size=14, bold=True)
    h3 = Font(size=13, bold=True)
    regular = Font(size=14)
    detailed = Font(size=13)


class Alignments:
    center = Alignment(vertical='center', horizontal='center')
    left = Alignment(vertical='center', horizontal='left', indent=1)


class Colors:
    success = PatternFill(start_color='abdeb9', fill_type='solid')
    error = PatternFill(start_color='F9A4A4', fill_type='solid')
    info = PatternFill(start_color='A4ABF9', fill_type='solid')

    detailed_success = PatternFill(start_color='C1F0C5', fill_type='solid')
    detailed_error = PatternFill(start_color='F0C1C1', fill_type='solid')
    detailed_info = PatternFill(start_color='BABFF8', fill_type='solid')


class Styles:
    test_success = NamedStyle('test_success')
    test_success.font = Fonts.regular
    test_success.alignment = Alignments.center
    test_success.fill = Colors.success

    test_error = NamedStyle('test_error')
    test_error.font = Fonts.regular
    test_error.alignment = Alignments.center
    test_error.fill = Colors.error

    detailed_header = NamedStyle('detailed_header')
    detailed_header.font = Fonts.h3
    detailed_header.alignment = Alignments.center
    detailed_header.fill = Colors.detailed_info

    detailed_success = NamedStyle("detailed_success")
    detailed_success.font = Fonts.detailed
    detailed_success.alignment = Alignments.center
    detailed_success.fill = Colors.detailed_success

    detailed_error = NamedStyle("detailed_error")
    detailed_error.font = Fonts.detailed
    detailed_error.alignment = Alignments.center
    detailed_error.fill = Colors.detailed_error


class Config:
    delete_previous = True,
    base_column_width = 25.0,
    base_row_height = 37.5,
    detailed_row_height = 18,
