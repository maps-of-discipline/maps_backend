import json
from io import BytesIO
from math import ceil
from pprint import pprint

from auth.logic import approved_required, login_required
from auth.models import Users
from cabinet.utils.excel_tools import create_excel_lessons_report, create_performance_report
from maps.logic.tools import timeit, LineTimer
from maps.models import D_ControlType, SprDiscipline, db, AupData, AupInfo, SprFaculty, Department, Groups, SprFormEducation
from cabinet.models import DisciplineTable, StudyGroups, Topics, Students, Grade, GradeTable, GradeType, GradeColumn, SprBells, SprPlace, TutorsOrder, TutorsOrderRow, Tutors

from flask import Blueprint, make_response, jsonify, request, send_from_directory, send_file
from cabinet.utils.serialize import serialize
from cabinet.lib.generate_discipline_table import generate_discipline_table
import datetime
from dateutil.parser import parse
from docxtpl import DocxTemplate
import functools 
import uuid


from openpyxl import load_workbook
import os

# regexp
import re

import requests

from itertools import groupby

cabinet = Blueprint('cabinet', __name__)

# Тестовый роут
@cabinet.route('/ping')
def test():
    return make_response('pong', 200)

### Задания
from time import time

# Получение данных таблицы "Задания"
@cabinet.route('/lessons', methods=['GET'])
@login_required
@approved_required
def getLessons():
    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')
    semester = request.args.get('semester')

    group = StudyGroups.query.filter_by(title=group_num).first()
    aup_info: AupInfo = AupInfo.query.filter_by(num_aup=num_aup).first()
    discipline_table = DisciplineTable.query.filter_by(id_aup=aup_info.id_aup, id_unique_discipline=id_discipline, study_group_id=group.id, semester=semester).first()

    control_types = getControlTypes(aup_info.id_aup, id_discipline, semester)
    amounts = [el['amount'] for el in control_types]
    sum_hours = sum(amounts)
    row_count = ceil(sum_hours / 2)

    if not discipline_table:
        res = generate_discipline_table(num_aup, id_discipline, group_num, semester, row_count)
        if 'error' in res:
            return jsonify(res['data'])
        else:
            discipline_table = res['data']

    response_data = {
        'topics': [],
        'discipline_table_id': discipline_table.id,
        'title': None
    }

    spr_discipline = SprDiscipline.query.filter(SprDiscipline.id == id_discipline).first()
    response_data['title'] = spr_discipline.title

    topics: Topics = Topics.query.filter(Topics.discipline_table_id == discipline_table.id).all()
    response_data['topics'] = serialize(topics)

    places = SprPlace.query.all()
    response_data['places'] = serialize(places)

    response_data['control_types'] = control_types


    return jsonify(response_data)



@cabinet.route('/lessons-excel', methods=['GET'])
# @login_required
# @approved_required
def get_lessons_as_xlsx():
    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')
    semester = request.args.get('semester')

    group = StudyGroups.query.filter_by(title=group_num).first()
    aup_info: AupInfo = AupInfo.query.filter_by(num_aup=num_aup).first()

    discipline_table: DisciplineTable = DisciplineTable.query.filter_by(
        id_aup=aup_info.id_aup,
        id_unique_discipline=id_discipline,
        study_group_id=group.id,
        semester=semester
    ).first()

    if not discipline_table:
        return jsonify({"error": "Not Found"}), 404

    file = create_excel_lessons_report(discipline_table.topics)
    file.seek(0)
    return send_file(
        file,
        download_name=f'{SprDiscipline.query.get(id_discipline).title}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



# Создание строки "Задания"
@cabinet.route('/lesson', methods=['POST'])
@login_required
@approved_required
def createLesson():
    data = request.get_json()

    if not data:
        return make_response('Отсутствует поле "lesson"', 400)

    new_lesson = Topics(
        topic=data['topic'],
        chapter=data['chapter'],
        id_type_control=data['id_type_control'],
        task_link=data['task_link'],
        task_link_name=data['task_link_name'],
        completed_task_link=data['completed_task_link'],
        completed_task_link_name=data['completed_task_link_name'],
        discipline_table_id=data['discipline_table_id'],
        study_group_id=data['study_group_id'],
        spr_place_id = data['spr_place_id'],
        place_note = data['place_note'],
        note = data['note'],
    )

    db.session.add(new_lesson)
    db.session.commit()

    res = serialize(new_lesson)
    return make_response(jsonify(res))

# Редактирование строки "Задания"
@cabinet.route('/lesson', methods=['PATCH'])
@login_required
@approved_required
def editLesson():
    data = request.get_json()

    if 'lesson' not in data:
        return make_response('Отсутствует поле "lesson"', 401)

    topic = Topics.query.filter(Topics.id == data['lesson']['id']).first()

    topic.chapter = data['lesson']['chapter']
    topic.topic = data['lesson']['topic']
    topic.task_link = data['lesson']['task_link']
    topic.task_link_name = data['lesson']['task_link_name']
    topic.completed_task_link = data['lesson']['completed_task_link']
    topic.completed_task_link_name = data['lesson']['completed_task_link_name']
    topic.id_type_control = data['lesson']['id_type_control']
    topic.date_task_finish_include = data['lesson']['date_task_finish_include']
    topic.spr_bells_id = data['lesson']['spr_bells_id']
    topic.spr_place_id = data['lesson']['spr_place_id']
    topic.place_note = data['lesson']['place_note']
    topic.note = data['lesson']['note']

    if data['lesson']['date'] == None:
        topic.date = None
    else:
        topic.date = parse(data['lesson']['date']).astimezone(datetime.timezone(datetime.timedelta(hours=3)))
        """ date_obj = datetime.datetime.strptime(data['lesson']['date'], r'%Y-%m-%dT%H:%M:%S.%f%z')
        topic.date = date_obj.astimezone(datetime.timezone(datetime.timedelta(hours=3))) """


    if data['lesson']['date_task_finish'] == None:
        topic.date_task_finish = None
    else:
        topic.date_task_finish = parse(data['lesson']['date_task_finish']).astimezone(datetime.timezone(datetime.timedelta(hours=3)))
        """ date_obj = datetime.datetime.strptime(data['lesson']['date_task_finish'], r'%Y-%m-%dT%H:%M:%S.%f%z')
        topic.date_task_finish = date_obj.astimezone(datetime.timezone(datetime.timedelta(hours=3))) """

    topic.lesson_order = data['lesson']['lesson_order']

    db.session.add(topic)
    db.session.commit()

    if len(data['lesson']['task_link_name']) != 0:
        grade_types = GradeType.query.filter_by(discipline_table_id=topic.discipline_table.id, type='tasks').all()

        for grade_type in grade_types:
            grade_column = GradeColumn.query.filter_by(topic_id=topic.id, discipline_table_id=topic.discipline_table.id, grade_type_id=grade_type.id).first()

            if not grade_column:
                new_grade_column = GradeColumn(discipline_table_id=topic.discipline_table.id, grade_type_id=grade_type.id, topic_id=topic.id)
                db.session.add(new_grade_column)
    
    if topic.date:
        types = ['activity', 'attendance']

        for type in types: 
            grade_types = GradeType.query.filter_by(discipline_table_id=topic.discipline_table.id, type=type).all()

            for grade_type in grade_types:
                grade_column = GradeColumn.query.filter_by(topic_id=topic.id, discipline_table_id=topic.discipline_table.id, grade_type_id=grade_type.id).first()
                
                if not grade_column:
                    new_grade_column = GradeColumn(discipline_table_id=topic.discipline_table.id, grade_type_id=grade_type.id, topic_id=topic.id)
                    db.session.add(new_grade_column)
    
    db.session.commit()

    res = serialize(topic)

    return make_response(jsonify(res))

# Удаление строки "Задания"
@cabinet.route('/lesson', methods=['DELETE'])
@login_required
@approved_required
def deleteLesson():
    id = request.args.get('id')

    if not id:
        return make_response('Отсутствует "id"', 400)

    lesson = Topics.query.filter_by(id=id).first()

    res = serialize(lesson)

    db.session.delete(lesson)
    db.session.commit()

    return make_response(jsonify(res), 200)

###

### Оценки

# Метод для обновления списка студентов по группе в базе данных
def bulkInsertStudentsByGroup(group):
    from app import app
    payload = {
        'getStudents': '',
        'group': group,
        'token': app.config.get('LK_TOKEN')
    }
    print(app.config.get('LK_TOKEN'))

    res = requests.get(app.config.get('LK_URL'), params=payload)
    data = res.json()
    data_students = data['items']

    bulk_students = []

    group_obj = StudyGroups.query.filter_by(title=group).first()

    group_obj_s = serialize(group_obj)

    for student in data_students:
        lk_id = student['id']
        is_exist = bool(Students.query.filter_by(lk_id=lk_id).first())

        if not is_exist:
            bulk_students.append(Students(name=student['fio'], study_group_id=group_obj_s['id'], lk_id=lk_id))

    db.session.bulk_save_objects(bulk_students)
    db.session.commit()

    return bulk_students

# Получение данных таблицы "Успеваемость"
@cabinet.route('/grades', methods=['GET'])
@login_required
@approved_required
def getGrades():
    if 'aup' not in request.args:
        return make_response('Отсутствует параметр "aup"', 400)

    if 'id' not in request.args:
        return make_response('Отсутствует параметр "id"', 400)

    if 'group' not in request.args:
        return make_response('Отсутствует параметр "group"', 400)
    
    if 'semester' not in request.args:
        return make_response('Отсутствует параметр "semester"', 400)

    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')
    semester = request.args.get('semester')

    aup_info: AupInfo = AupInfo.query.filter(AupInfo.num_aup == num_aup).first()
    if not aup_info:
        return jsonify({'error': 'Данный АУП отсутствует.'})

    discipline_is_exist = AupData.query.filter(AupData.id_discipline == id_discipline).first()
    if not discipline_is_exist:
        return jsonify({'error': 'Дисциплина отсутствует в АУП.'})

    group = StudyGroups.query.filter(StudyGroups.title == group_num).first()
    if not group:
        return jsonify({'error': 'Данная группа отсутствует.'})

    discipline_table = DisciplineTable.query.filter_by(id_aup=aup_info.id_aup, id_unique_discipline=id_discipline, study_group_id=group.id, semester=semester).first()

    if not discipline_table:
        return jsonify({
            'is_not_exist': True,
            'message': 'Таблица успеваемости отсутствует'
        })

    grade_types = GradeType.query.filter_by(discipline_table_id=discipline_table.id).all()

    """ bulkInsertStudentsByGroup(group.title) """

    students = Students.query.filter(Students.study_group_id == group.id).all()
    # students = serialize(students)

    rows = []
    for student in students:
        student: Students
        grades = [grade.to_dict(rules=['-student', '-grade_column']) for grade in student.grades]

        values = {}
        for grade in grades:
            values[grade['grade_column_id']] = grade['value']

        rows.append({
            'id': student.id,
            'name': student.name,
            'values': values
        })

    columns = GradeColumn.query.filter_by(discipline_table_id=discipline_table.id).all()

    return jsonify({
        'gradeTypes': serialize(grade_types),
        'disciplineTableId': discipline_table.id,
        'columns': [col.to_dict(rules=['-grades']) for col in columns],
        'rows': rows
    })

# Редактирование оценки
@cabinet.route('/grade', methods=['PATCH'])
@login_required
@approved_required
def updateGrade():
    data = request.get_json()

    grade = Grade.query.filter_by(student_id=data['student_id'],
                                  grade_column_id=data['grade_column_id']).first()

    if data['value'] == None:
        data['value'] = 0

    if not grade:
        grade = Grade(value=data['value'], student_id=data['student_id'],
                      grade_column_id=data['grade_column_id'])
        db.session.add(grade)
    else:
        grade.value = data['value']

    db.session.commit()

    res = grade.to_dict(rules=['-grade_column.grades'])

    return jsonify(res)

###

### Столбцы оценок

@cabinet.route('/grade-column-visible', methods=['PATCH'])
@login_required
@approved_required
def updateGradeColumn():
    data = request.get_json()

    id = data['id']
    visible_column_ids = data['visible']
    hidden_column_ids = data['hidden']

    visible_columns = GradeColumn.query.filter(GradeColumn.id.in_(visible_column_ids)).all()
    hidden_columns = GradeColumn.query.filter(GradeColumn.id.in_(hidden_column_ids)).all()

    
    for col in visible_columns:
        col.hidden = False

    for col in hidden_columns:
        col.hidden = True

    db.session.bulk_save_objects(visible_columns)
    db.session.bulk_save_objects(hidden_columns)
    db.session.commit()

    res = []
    for col in visible_columns:
        res.append(col.to_dict(rules=['-grades']))

    for col in hidden_columns:
        res.append(col.to_dict(rules=['-grades']))

    return jsonify(res)

### Виды оценивания

# Получение видов оценивания
@cabinet.route('/grade-type', methods=['GET'])
@login_required
@approved_required
def getGradeType():
    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')

    aup_info: AupInfo = AupInfo.query.filter(AupInfo.num_aup == num_aup).first()
    if not aup_info:
        return jsonify({'error': 'Данный АУП отсутствует.'})

    group = StudyGroups.query.filter(StudyGroups.title == group_num).first()
    if not group:
        return jsonify({'error': 'Данная группа отсутствует.'})

    discipline_table = DisciplineTable.query.filter_by(id_aup=aup_info.id_aup, id_unique_discipline=id_discipline,
                                             study_group_id=group.id).first()
    grade_types = GradeType.query.filter_by(discipline_table_id=discipline_table.id).all()

    return jsonify(serialize(grade_types))

# Создание нового вида оценивания
@cabinet.route('/grade-type', methods=['POST'])
@login_required
@approved_required
def createGradeType():
    data = request.get_json()

    grade_type = GradeType(name=data['name'], discipline_table_id=data['table_id'], type=data['type'], is_custom=True)
    db.session.add(grade_type)
    db.session.commit()
    
    discipline_table = DisciplineTable.query.filter_by(id=data['table_id']).first()
    
    need_add_cols = []
    for topic in discipline_table.topics:
        if grade_type.type == 'attendance' and topic.date:
            need_add_cols.append(GradeColumn(discipline_table_id=data['table_id'], grade_type_id=grade_type.id, topic_id=topic.id))
        if grade_type.type == 'tasks' and topic.task_link_name:
            need_add_cols.append(GradeColumn(discipline_table_id=data['table_id'], grade_type_id=grade_type.id, topic_id=topic.id))

    print(need_add_cols)

    db.session.bulk_save_objects(need_add_cols)
    db.session.commit()

    cols = GradeColumn.query.filter_by(discipline_table_id=data['table_id'], grade_type_id=grade_type.id).all()

    return jsonify({
        'grade_type': serialize(grade_type),
        'columns': serialize(cols),
    })

# Редактирование видов оценивания
@cabinet.route('/grade-type', methods=['PATCH'])
@login_required
@approved_required
def updateGradeType():
    data = request.get_json()

    grade_type = GradeType.query.filter_by(id=data['id']).first()

    grade_type.archived = data['archived']
    grade_type.name = data['name']
    grade_type.min_grade = data['min_grade']
    grade_type.max_grade = data['max_grade']    
    grade_type.binary = data['binary']
    grade_type.weight_grade = data['weight_grade']

    db.session.commit()

    res = serialize(grade_type)

    return jsonify(res)

@cabinet.route('/grade-type', methods=['DELETE'])
@login_required
@approved_required
def deleteGradeType():
    id = request.args.get('id')

    grade_type = GradeType.query.filter_by(id=id).first()

    db.session.delete(grade_type)
    db.session.commit()

    return jsonify(1)

###

### Пользователи

# Получение данных о пользователе
@cabinet.route('/user', methods=['GET'])
@login_required
@approved_required
def getUser():
    token = request.args.get('token')

    if not token:
        return make_response('Отсутствует "token"', 400)

    payload = {'getUser': '', 'token': token}

    from app import app
    res = requests.get(app.config.get('LK_URL'), params=payload)

    user = res.json()['user']

    user_db = Users.query.filter_by(lk_id=user['id']).first()

    roles = user_db.roles
    print(roles)

    permissions = {}
    for role in roles:
        if (role.name_role == 'student'):
            student = Students.query.filter_by(lk_id=user_db.lk_id).first()

            if not student:
                continue

            study_group = StudyGroups.query.filter_by(id=student.study_group_id).first()

            if (study_group):
                permissions['student'] = {
                    'aup': study_group.num_aup,
                    'group': {
                        'id': study_group.id,
                        'title': study_group.title,
                    },
                }
        # elif (role.name_role == 'admin'):
        #     # Add admin-specific permissions
        #     permissions['admin'] = {
        #         'isAdmin': True,
        #         # Add any other admin-specific permissions needed
        #         'canAccessAdminPanel': True,
        #         'canManageUsers': True,
        #     }

    return jsonify({
        'name': user['name'],
        'surname': user['surname'],
        'avatar': user['avatar'],
        'permissions': permissions
    })

# Получение пользователей системы
@cabinet.route('/lk-users', methods=['GET'])
@login_required
@approved_required
def getLKUsers():
    users = Users.query.filter_by(auth_type='lk').all()

    res = []
    for user in users:
        roles = []
        for role in user.roles:
            roles.append({
                "id_role": role.id_role,
                "name_role": role.name_role
            })

        res.append({
            'id_user': user.id_user,
            'name': user.name,
            'approved_lk': user.approved_lk,
            'request_approve_date': user.request_approve_date,
            'roles': roles,
        })

    return jsonify(res)

# Выдача доступа для пользователя
@cabinet.route('/approve-user', methods=['PATCH'])
@login_required
@approved_required
def updateApproveUser():
    data = request.get_json()

    user: Users = Users.query.filter_by(id_user=data['id_user']).first()
    user.approved_lk = data['value']

    db.session.commit()

    return jsonify(True)

### 

### Группы

# Загрузка файла выгрузки из 1С "Соответствие групп и учебных планов"
# и формирование на его основе таблицы в базе данных
@cabinet.route('/groups', methods=['PATCH'])
@login_required
@approved_required
def uploadGroups():
    files = request.files.getlist("file")
    file = files[0]

    from app import app

    path = os.path.join(app.static_folder, 'temp', file.filename)
    file.save(path)

    wb = load_workbook(file)
    sheet = wb.active

    study_groups = []

    for i in range(1, sheet.max_row + 1):
        group_cell = sheet.cell(row=i, column=2)
        aup_cell = sheet.cell(row=i, column=3)

        group_name = group_cell.value
        num_aup_regexp = re.search(r'\d{9}', aup_cell.value)

        if not num_aup_regexp == None:
            num_aup = num_aup_regexp[0]

            study_groups.append(StudyGroups(title=group_name, num_aup=num_aup))

    db.session.bulk_save_objects(study_groups)
    db.session.commit()

    res = serialize(study_groups)

    return jsonify(res)


# Получение списка доступных групп
@cabinet.route('/groups', methods=['GET'])
@login_required
@approved_required
def getGroups():
    num_aup = request.args.get('aup')

    groups = None
    if num_aup:
        groups = StudyGroups.query.filter_by(num_aup=num_aup).all()
    else:
        groups = StudyGroups.query.all()

    res = serialize(groups)

    return jsonify(res)
###

### Звонки

# Получение расписания звонков
@cabinet.route('/bells', methods=['GET'])
@login_required
@approved_required
def getBells():
    bells = SprBells.query.all()
    return jsonify(serialize(bells))

# Обновление звонков
@cabinet.route('/bells', methods=['PATCH'])
@login_required
@approved_required
def updateBells():
    data = request.get_json()

    db.session.bulk_update_mappings(SprBells, data)
    db.session.commit()

    return jsonify('ok')

###

### Отчеты

# Получить отчет по дисциплине
@cabinet.route('/report-by-discipline', methods=['GET'])
@login_required
@approved_required
def getReportByDiscipline():
    id_discipline = request.args.get('id_discipline')

    rpds: DisciplineTable = DisciplineTable.query.filter(DisciplineTable.id_unique_discipline == id_discipline).all()

    grades = {
        2: 0,
        3: 0,
        4: 0,
        5: 0,
    }

    for rpd in rpds:
        for topic in rpd.topics:
            for grade in topic.grades:
                grades[grade.value] += 1

    res = grades

    return jsonify(res)

# Получение отчета по группе
@cabinet.route('/report', methods=['GET'])
@login_required
@approved_required
def getReport():
    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')

    aup_info: AupInfo = AupInfo.query.filter(AupInfo.num_aup == num_aup).first()
    if not aup_info:
        return jsonify({'error': 'Данный АУП отсутствует.'})

    group = StudyGroups.query.filter(StudyGroups.title == group_num).first()
    if not group:
        return jsonify({'error': 'Данная группа отсутствует.'})

    discipline_table = DisciplineTable.query.filter_by(id_aup=aup_info.id_aup, id_unique_discipline=id_discipline,
                                             study_group_id=group.id).first()

    grade_types = GradeType.query.filter_by(discipline_table_id=discipline_table.id).all()

    gcs = GradeColumn.query.filter_by(discipline_table_id=discipline_table.id).all()

    grades = []
    for column in gcs:
        for grade in column.grades:
            grades.append(grade)


    grades = [grade.to_dict(rules=['-grade_column.grades']) for grade in grades]

    grouped_grades_by_students = groupby(sorted(grades, key=lambda x: x['student_id']), key=lambda x: x['student_id'])

    result = {}

    for key, grades_student in grouped_grades_by_students:
        grouped_by_grade_type = groupby(
            sorted(grades_student, key=lambda x: x['grade_column']['grade_type_id']),
            key=lambda x: x['grade_column']['grade_type_id']
        )

        if key not in result:
            result[key] = dict()

        result[key]['categories'] = {}
        for grade_type in grade_types:
            result[key]['categories'][grade_type.id] = {
                'grade_type_id': grade_type.id,
                'name': grade_type.name,
                'value': 0,
            }

        for key_cols, cols in grouped_by_grade_type:
            if 'categories' not in result[key]:
                result[key]['categories'] = dict()

            list_cols = list(cols)

            if 'name' not in result[key]:
                result[key]['name'] = list_cols[0]['student']['name']

            grade_type_id = list_cols[0]['grade_column']['grade_type_id']
            for col in list_cols:
                if isinstance(col['value'], int):
                    result[key]['categories'][grade_type_id]['value'] = result[key]['categories'][grade_type_id]['value'] + col['value']

    return jsonify({
        'rating_chart': result,
    })

###

### Остальное

# Получение всех нагрузок дисциплины
def getControlTypes(id_aup, id_unique_discipline, semester):
    diciplines = AupData.query.filter_by(id_aup=id_aup, id_discipline=id_unique_discipline, id_period=semester).all()
    serialized_diciplines = [discipline.to_dict() for discipline in diciplines]

    control_types = {}
    control_type_q = D_ControlType.query.all()
    for row in control_type_q:
        control_types[row.id] = {
            'title': row.title,
            'shortname': row.shortname
        }

    # Преобразует все строки из выгрузки в список нагрузок на дисциплине
    def mapDisciplinesToControlType(dicipline):
        id = dicipline['id_type_control']

        amount = dicipline['amount'] / 100
        if dicipline['id_edizm'] == 2:
            amount = amount * 54

        return {
            'id_type_control': id,
            'name': control_types[id]['title'],
            'shortname': control_types[id]['shortname'],
            'amount': amount,
            'id_period': dicipline['id_period']
        }

    control_types = list(map(mapDisciplinesToControlType, serialized_diciplines))
    filtered_control_types = filter(lambda ct: ct['amount'] != 0, control_types)

    return list(filtered_control_types)

# Получение всех нагрузок дисциплины
@cabinet.route('/place', methods=['GET'])
@login_required
@approved_required
def getPlace():
    faculties = SprPlace.query.all()
    return jsonify(serialize(faculties))

# Получение данных об учебном плане по поиску
@cabinet.route('/aup', methods=['GET'])
@login_required
@approved_required
def getAup():
    num_aup = request.args.get('aup')
    search = request.args.get('search')

    res = None
    if search:
        found: list[AupInfo] = AupInfo.query.filter(AupInfo.file.like("%" + search + "%")).all()
        res = []
        for aup in found:
            res.append({
                'num_aup': aup.num_aup,
                'id': aup.id_aup,
                'title': aup.name_op.name_spec,
                'year_beg': aup.year_beg,
            })
    else:
        found: AupInfo = AupInfo.query.filter_by(num_aup=num_aup).first()
        if found:
            res = {
                'num_aup': found.num_aup,
                'id': found.id_aup,
                'title': found.name_op.name_spec,
                'semesters': [1, 2, 3, 4, 5, 6, 7, 8]
            }
        else:
            res = {
                'error': 'АУП не найден',
                'semesters': []
            }

    return jsonify(res)

# Получение списка дисциплин учебного плана
@cabinet.route('/disciplines', methods=['GET'])
@login_required
@approved_required
def disciplines():
    q_num_aup = request.args.get('aup')
    aup = AupInfo.query.filter(AupInfo.num_aup == q_num_aup).first()
    disciplines = AupData.query.filter(AupData.id_aup == aup.id_aup).all()
    
    res = list({(el.id_discipline, el._discipline) for el in disciplines})
    res = [dict(id=el[0], title=el[1]) for el in res]
    return jsonify(res)

# Получение списка дисциплин учебного плана (вариант для студентов)
@cabinet.route('/disciplines-new', methods=['GET'])
@login_required
@approved_required
def disciplinesNew():
    num_aup = request.args.get('aup')

    aup_info = AupInfo.query.filter(AupInfo.num_aup == num_aup).first()
    aup_data = AupData.query.filter_by(id_aup=aup_info.id_aup).order_by(AupData.shifr, AupData._discipline,
                                                                        AupData.id_period).all()

    disciplines_items = {}
    flag = ""

    group_color_mapper  = {el.id_group: el.color for el in Groups.query.all()}

    for i, item in enumerate(aup_data):
        if flag != item.discipline + str(item.id_period):
            flag = item.discipline + str(item.id_period)

            d = dict()

            d["id"] = item.id_discipline
            d["name"] = item.unique_discipline.title
            d["num_row"] = item.num_row
            d["color"] = group_color_mapper[item.id_group]

            if (item.id_period in disciplines_items):
                disciplines_items[item.id_period].append(d)
            else:
                disciplines_items[item.id_period] = [d]

    return jsonify(disciplines_items)


# Получение информации о дисциплине
@cabinet.route('/discipline', methods=['GET'])
@login_required
@approved_required
def discipline():
    id = request.args.get('id')
    
    spr_discipline = SprDiscipline.query.filter_by(id=id).first()

    return jsonify(serialize(spr_discipline))


# Скачать распоряжение тьюторов
@cabinet.route('/download-tutor-order', methods=['POST'])
@login_required
@approved_required
def downloadTutorOrder():
    body = request.get_json()

    template = {
        "date": "01.09.2023",
        "order": "10-Р",
        "year": "2023/2024",
        "faculty": "факультет информационных технологий",
        "form_education": "очная",
        "body": body,
        "need_report": "раз в месяц",
        "need_report_day": "первый понедельник",
        "provide_person": {
            "jobtitle": "заместитель декана по общим вопросам",
            "name": "В.М. Черновой"
        },
        "signer": {
            "name": "Д.Г. Демидов"
        },
        "executor": "Олейникова Е.В., тел: 1704"
    }

    docx = DocxTemplate('static/docx_templates/tutor_template.docx')


    docx.render(template)
    docx.save('static/docx_templates/tutor_template_res.docx')

    return send_from_directory('static/docx_templates', 'tutor_template_res.docx', as_attachment=True)

# Получение списка факультетов
@cabinet.route('/faculties', methods=['GET'])
@login_required
@approved_required
def getFaculties():
    faculties = SprFaculty.query.all()

    res = [faculty.to_dict(rules=['-departments.faculty', '-departments.tbl_aups']) for faculty in faculties]

    return jsonify(res)

# Получение списка кафедр
@cabinet.route('/departments', methods=['GET'])
@login_required
@approved_required
def getDepartments():
    departments = Department.query.all()
    return jsonify(serialize(departments))

# Получение списка сотрудников
@cabinet.route('/staff', methods=['GET'])
@login_required
@approved_required
def getStaff():
    name_department = request.args.get('division')

    department = Department.query.filter_by(name_department=name_department).first()
    
    print(department, name_department)


    from app import app
    payload = {
        'getStaff': '',
        'division': name_department,
        'token': app.config.get('LK_TOKEN')
    }

    res = requests.get(app.config.get('LK_URL'), params=payload)
    data = res.json()
    staffs = data['items']

    has_tutors_map_lk_ids = [tutor.lk_id for tutor in Tutors.query.filter_by(id_department=department.id_department).all()]

    need_add = []
    for staff in staffs:
        converted_staff = Tutors(
            lk_id=staff['id'],
            name=staff['fio'],
            id_department=department.id_department
        )

        if (int(converted_staff.lk_id) not in has_tutors_map_lk_ids):
            need_add.append(converted_staff)
    
    db.session.bulk_save_objects(need_add)
    db.session.commit()

    tutors = Tutors.query.filter_by(id_department=department.id_department).all()

    return jsonify(serialize(tutors))

def getDisciplinesByTeacher(fullname):
    payload = {
        'getScheduleTeacher': '',
        'fio': fullname,
        "token": "fVuzytCC7l7g31Rj0%2FjUR4HJ4gvzRZGLwjMvp6wmM4b91x%2FmFKwVJffBQA8DN0XpdvhClndCo5wC7Ii6HHqiQvCLMvnS5%2BzwFg1t%2BzVEIuqI0ddC52M43dyJAiK70En911P8cTRIt3CnCLqVO8sAIjyE%2Bnrpu9gXx%2BvrhSDSiW4%3D",
        'session': 1,
    }
    
    from app import app
    res = requests.get(app.config.get('LK_URL'), params=payload)
    data = json.loads(res.text)

    data = dict(filter(lambda x: bool(x[1]), data.items()))
    disciplines = {}
    for value in data.values():
        for discipline in value:
            discipline_name = discipline['name'].split('(')[0]
            if discipline_name not in disciplines:
                disciplines.update({discipline_name: {*discipline['groups'].split(', ')}})
            else:
                disciplines[discipline_name] |= {*discipline['groups'].split(', ')}

    disciplines = [{"name": key, "groups": list(value)} for key, value in disciplines.items()]
    return jsonify(disciplines), 200


@cabinet.route('/discipline-by-aup')
def get_discipline_by_aup():
    id_aup = request.args['id_aup']
    id_discipline = request.args['id_discipline']

    aup_data = AupData.query.filter_by(id_aup=id_aup, id_discipline=id_discipline).all()

    if not aup_data:
        return jsonify({"error": "Not found"}), 400

    res = {"title": aup_data[0].discipline, "periods": list({el.id_period for el in aup_data})}

    return jsonify(res)

# Получение распоряжения
@cabinet.route('/tutor-orders', methods=['GET'])
@login_required
@approved_required
def tutorOrders():
    faculty_id = request.args.get('id')

    tutor_orders = TutorsOrder.query.filter_by(faculty_id=faculty_id).all()
    
    res = []
    for tutor_order in tutor_orders:
        tutor_order_rows = TutorsOrderRow.query.filter_by(tutors_order_id=tutor_order.id).all()
        tutor_order_rows = [tutor_order_row.to_dict(rules=['-department.tbl_aups', '-department.faculty']) for tutor_order_row in tutor_order_rows]

        exist_department_tutors= []
        grouped_tutor_order_rows = {}
        for row in tutor_order_rows:
            if row['tutor']['lk_id'] in exist_department_tutors:
                continue

            if row['department_id'] not in grouped_tutor_order_rows:
                grouped_tutor_order_rows[row['department_id']] = {
                    'id_department': row['department_id'],
                    'name_department': row['department']['name_department'],
                    'rows': []
                }

            tutor_rows = TutorsOrderRow.query.filter_by(tutor_id=row['tutor']['id'], department_id=row['department']['id_department']).all()
            study_groups = [tr.study_group.to_dict() for tr in tutor_rows]

            grouped_tutor_order_rows[row['department_id']]['rows'].append({
                'id': uuid.uuid4(),
                'tutor': row['tutor'],
                'study_groups': study_groups,
            })

            exist_department_tutors.append(row['tutor']['lk_id'])

        res.append({
            'meta': serialize(tutor_order),
            'body': grouped_tutor_order_rows,
        })

    return jsonify(res)

# Редактирование распоряжение тьюторов
@cabinet.route('/tutor-order', methods=['PATCH'])
@login_required
@approved_required
def editTutorOrder():
    data = request.get_json()
    print(data)

    id_order = data['id']
    meta = data['meta']
    body = data['body']

    
    """ tutors = []
    tutors_ids = []
    groups_by_lk_id = {}
    for department in body:
        rows = department['rows']

        tutors_department = []
        tutors_department_ids = []
        for row in rows:
            tutor = Tutors(
                name= row['tutor']['fio'], 
                post=row['tutor']['post'], 
                id_department=department['id_department'], 
                lk_id=int(row['tutor']['id'])
            )

            tutors_department.append(tutor)
            tutors_department_ids.append(tutor.lk_id)
            groups_by_lk_id[tutor.lk_id] = row['study_groups']

        tutors.extend(tutors_department)
        tutors_ids.extend(tutors_department_ids)


    found_tutors = Tutors.query.filter(Tutors.lk_id.in_(tutors_ids)).all()
    exist_tutors_ids = [ exist_tutor.lk_id for exist_tutor in found_tutors ]

    need_add_tutors = []
    for tutor in tutors:
        if (tutor.lk_id in exist_tutors_ids):
            pass
        else:
            need_add_tutors.append(tutor)


    db.session.bulk_save_objects(need_add_tutors)
    db.session.commit() 

    exist_tutors = Tutors.query.filter(Tutors.lk_id.in_(tutors_ids)).all()

    for tutor in exist_tutors:
        study_group_ids = [ group['id'] for group in groups_by_lk_id[tutor.lk_id] ]
        StudyGroups.query.filter(StudyGroups.id.in_(study_group_ids)).update({ StudyGroups.tutor_id: tutor.id }) """

    tutor_order = TutorsOrder.query.filter_by(id=id_order).first()

    tutor_order.date = parse(meta['date']).astimezone(datetime.timezone(datetime.timedelta(hours=3)))
    tutor_order.num_order = meta['order']
    tutor_order.year = meta['year']
    tutor_order.executor = meta['executor']
    tutor_order.signer = meta['signer']

    tutors_order_rows = TutorsOrderRow.query.filter_by(tutors_order_id=id_order).all()

    TutorsOrderRow.query.filter_by(tutors_order_id=tutor_order.id).delete(synchronize_session=False)

    new_tutors_order_rows = []
    for department in body:
        for row in department['rows']:
            for study_group in row['study_groups']:
                new_tutors_order_rows.append(TutorsOrderRow(
                    tutors_order_id=tutor_order.id,
                    department_id=department['id_department'],
                    study_group_id=study_group['id'],
                    tutor_id=row['tutor']['id'],
                ))

    print(new_tutors_order_rows)
    db.session.bulk_save_objects(new_tutors_order_rows)


    """tutors_order_rows_map = [ row.id for row in tutors_order_rows ]

     need_add_row = []
    for department in body:
        rows = department['rows']

        for row in rows:
            if row['id'] not in tutors_order_rows_map:
                tutor = Tutors.query.filter_by(lk_id=row['tutor']['id']).all()

                need_add_row.append(TutorsOrderRow(
                    tutors_order_id=tutor_order.id,
                    department_id=department['id_department'],
                    tutor_id=tutor.id,
                )) """


    db.session.add(tutor_order)
    db.session.commit()

    
    res = {
        'meta': serialize(tutor_order),
        'body': [],
    }

    return jsonify(res)


# Получение формы обучения
@cabinet.route('/form-of-educations', methods=['GET'])
@login_required
@approved_required
def getFormOfEducations():
    form_of_educations = SprFormEducation.query.all()
    return jsonify(serialize(form_of_educations))


@cabinet.route('/performance-excel', methods=['GET'])
# @login_required
# @approved_required
def get_performance_excel_report():
    num_aup = request.args.get('aup')
    id_discipline = request.args.get('id')
    group_num = request.args.get('group')
    semester = request.args.get('semester')

    group = StudyGroups.query.filter_by(title=group_num).first()
    aup_info: AupInfo = AupInfo.query.filter_by(num_aup=num_aup).first()

    discipline_table: DisciplineTable = DisciplineTable.query.filter_by(
        id_aup=aup_info.id_aup,
        id_unique_discipline=id_discipline,
        study_group_id=group.id,
        semester=semester
    ).first()

    if not discipline_table:
        return jsonify({"error": "Not Found"}), 404

    file = create_performance_report(discipline_table, group)
    file.seek(0)
    return send_file(
        file,
        download_name=f'{SprDiscipline.query.get(id_discipline).title}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
