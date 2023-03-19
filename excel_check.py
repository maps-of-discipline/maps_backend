from collections import defaultdict
from math import ceil
import os
from openpyxl import load_workbook
from tools import get_maximum_rows, skiplist, sems
from models import SprStandard, SprVolumeDegreeZET


def check_smt1(file):  # проверка на целочисленность дисциплины №2
    wb = load_workbook(file, read_only=True)
    ws = wb['Лист2']
    max_row = get_maximum_rows(sheet_object=ws)
    d = defaultdict(list)
    for i in range(2, max_row):
        sem = ws['G'+str(i)].value
        disc = ws['F'+str(i)].value
        hours = ws['I'+str(i)].value
        ed_izm = ws['J'+str(i)].value
        record_type = ws['E'+str(i)].value
        block = ws['A'+str(i)].value

        if (len(list(filter(lambda x: x in disc, skiplist['discipline']))) > 0 or
                len(list(filter(lambda x: x in record_type, skiplist['record_type']))) > 0 or
                len(list(filter(lambda x: x in block, skiplist['record_type']))) > 0):
            continue
        if hours:
            
            try:
                hours = int(float(hours)*100)
            except:
                hours = int(float(hours.replace(",", "."))*100)
            if ed_izm == 'Недели': 
                hours = hours*54 # 54 = 1 ZET * 1.5 = 36 * 1.5

            d[sem].append([disc, hours])
    # Print semesters
    # for key, value in d.items():
    #     print()
    #     print("{0}: {1}".format(key,value))
    ret_arr = []
    for key, value in d.items():
        ddd = dict()
        for i in range(0, len(value)):
            if ddd.get(value[i][0]):
                # try:
                ddd[value[i][0]] += value[i][1]
                # except:
                #     ddd[value[i][0]] += value[i][1]
            else:
                # try:
                ddd[value[i][0]] = value[i][1]
                # except:
                #     ddd[value[i][0]] = value[i][1]
        for key1, value1 in ddd.items():
            if not (value1/3600).is_integer():
                ret_arr.append("{0}: {1} {2}".format(key, key1, value1/3600))
    return ret_arr


def check_smt(file):  # проверка на целочисленность дисциплины №1
    wb = load_workbook(file)
    ws = wb['Лист2']
    max_row = get_maximum_rows(sheet_object=ws)
    sumzet = 0.0
    arr_err = []
    table = []
    for i in range(2, max_row):
        period = ws['G'+str(i)].value
        discipline = ws['F'+str(i)].value
        zet = ws['K'+str(i)].value
        record_type = ws['E'+str(i)].value
        block = ws['A'+str(i)].value

        if (len(list(filter(lambda x: x in discipline, skiplist['discipline']))) > 0 or
                len(list(filter(lambda x: x in record_type, skiplist['record_type']))) > 0 or
                len(list(filter(lambda x: x in block, skiplist['record_type']))) > 0):
            continue

        if zet != None:
            # print(zet)
            try:
                zet = float(zet.replace(',', '.'))
            except:
                pass
            # print(zet)
            try:
                sumzet += zet
            except:
                sumzet += float(zet)
        else:
            continue
        period = period.split()[0]

        # словарь с данными ячейчи
        cell = {
            "discipline": discipline,
            "term": period,
            "zet": zet
        }

        # добаляем в таблицу недостоющее количество семестров для очередной записи в списке workload
        delta = len(table) - sems.index(period) - 1
        if delta < 0:
            for i in range(-delta):
                table.append([])

        # считаем сумму зет для каждой дисциплины, включая экзамены, лаб.занятия, лекции и т.д.
        for el in table[sems.index(period)]:
            if el['discipline'] == discipline:
                el['zet'] += zet
                break
        else:
            # если такой дисциплины нет, то добавляем в таблицу
            table[sems.index(period)].append(cell)
    # print(table)

    for semester in table:
        for disc in semester:
            print(disc)
            if not float(disc['zet']).is_integer():
                arr_err.append(disc['term'] + ' ' + 'семестр' +
                               ' ' + disc['discipline'] + ' ' + str(disc['zet']))
    print(arr_err)
    return arr_err


def check_empty_ceils(file):  # Проверка на пустые обязательные ячейки
    wb = load_workbook(file)
    ws = wb['Лист2']
    max_row = get_maximum_rows(sheet_object=ws)
    err_arr = []
    for letter in 'ABEFGHJ':
        for num in range(1, max_row + 1):
            if ws[letter+str(num)].value == None:
                err_arr.append(letter+str(num))
    # типа попытался реализовать проверку, чтобы если есть значение "Пустой элемент" или пустая ячейка и в "А" значение Факультатива, то пропускать дальше без ошибок
    # for num in range(1, max_row + 1): 
    #     # если в колонке 'C' пустое значение и в колонке 'A' с этим же номером это факультатиная дисциплина, то пропускать и добавлять в базу
    #     if (ws['C'+str(num)].value == None or ws['C'+str(num)].value == 'Пустой компонент') and len(list(filter(lambda x: x in ws['A'+str(num)].value, skiplist['record_type']))) == 0:
    #         err_arr.append(letter+str(num))
    # if err_arr == []:
    #     return True, err_arr
    # else:
    return err_arr


def layout_of_disciplines(file):  # Компоновка элективных дисциплин по семестрам
    wb = load_workbook(file)
    ws = wb['Лист2']
    max_row = get_maximum_rows(sheet_object=ws)
    for num in range(1, max_row + 1):
        if 'Элективные дисциплины' in ws['E'+str(num)].value:
            temp_num = ws['E'+str(num)].value
            temp_value = ws['F'+str(num)].value
            count = 0
            for i in range(num, max_row + 1):
                count += 1
                if ws['E'+str(i)].value == temp_num and ws['F'+str(i)].value != temp_value:
                    for j in range(1, count):
                        ws['F'+str(i-j)] = temp_value + ' / ' + \
                            ws['F'+str(i+(j-1))].value
                        ws['F'+str(i+(j-1))] = 'None'
                        ws['E'+str(i+(j-1))] = 'None'
    for num in range(1, max_row + 1):
        if ws['E'+str(num)].value == 'None':
            ws.delete_rows(num)
    for num in range(1, max_row + 1):
        if ws['E'+str(num)].value == 'None':
            ws.delete_rows(num)
    wb.save(file)


# Проверка, чтобы общая сумма ЗЕТ соответствовало норме (30 * кол-во семестров)
def check_full_zet_in_plan(file):
    wb = load_workbook(file)
    ws = wb['Лист2']
    column_semester = ws['G']
    column_hours = ws['I']
    column_edizm = ws['J']
    column_record_type = ws['E']
    column_discipline = ws['F']
    column_block = ws['A']
    temp_list = []
    for i in range(1, len(column_semester)):
        if column_semester[i].value not in temp_list:
            temp_list.append(column_semester[i].value)

    ws = wb['Лист1']
    program_code = ws['B6'].value
    standard = format_standard(ws['B9'].value)
    id_standard = SprStandard.query.filter_by(
        type_standard=standard).first().id_standard

    select = SprVolumeDegreeZET.query.filter_by(
        program_code=program_code, id_standard=id_standard).first()
    sum_normal = select.zet

    sum_zet = 0
    for i in range(1, len(column_hours)):
        if (column_hours[i].value is not None and (
                len(list(filter(lambda x: x in column_discipline[i].value, skiplist['discipline']))) == 0 and
                len(list(filter(lambda x: x in column_record_type[i].value, skiplist['record_type']))) == 0 and
                len(list(filter(lambda x: x in column_block[i].value, skiplist['record_type']))) == 0)):
            
            if column_edizm[i].value == 'Недели':
                try:
                    sum_zet += float(column_hours[i].value.replace(',', '.'))*54
                except:
                    sum_zet += float(column_hours[i].value)*54
            else:
                try:
                    sum_zet += float(column_hours[i].value.replace(',', '.'))
                except:
                    sum_zet += float(column_hours[i].value)
    sum_zet /= 36

    if abs(round(sum_zet, 2) - sum_zet) < 0.001:
        sum_zet = round(sum_zet, 2)

    # if sum_normal == sum_zet:
    #     return True, None, None
    # else:
    return sum_normal, sum_zet



def format_standard(standard):
    if standard == 'ФГОС3++' or standard == 'ФГОС ВО (3++)':
        standard = 'ФГОС ВО 3++'
    return standard


def excel_check(path, aup, options_check):
    return_err_arr = []
    ## ------------------------------------ ###
    ## Проверка на пустые ячейки ###
    err_arr = check_empty_ceils(path)
    if err_arr != []:
        errors = 'АУП: ' + aup + ' В документе не заполнены ячейки:' + ', '.join(err_arr)
        print(errors)
        return_err_arr.append(errors)
    ### ------------------------------------ ###

    # ### Проверка на целочисленность ЗЕТ у каждой дисциплины ###
    if options_check['enableCheckIntegrality'] == True:
        err_arr = check_smt1(path)
        if err_arr != []:
            errors = 'АУП: ' + aup + ' Ошибка при подсчете ЗЕТ:\n' + '\n'.join(err_arr)
            print(errors)
            return_err_arr.append(errors)
    # ### ------------------------------------ ###

    ### Компановка элективных курсов ###
    layout_of_disciplines(path)
    ### ---------------------------- ###

    # ### ------------------------------------ ###
    # ### Проверка, чтобы общая сумма ЗЕТ соответствовало норме (30 * кол-во семестров) ###
    if options_check['enableCheckSumMap'] == True:
        sum_normal, sum_zet = check_full_zet_in_plan(path)
        print(sum_normal, sum_zet)
        if sum_normal != sum_zet:
            errors = 'АУП: ' + aup + ' В выгрузке общая сумма ЗЕТ не соответствует норме. Норма {} ЗЕТ. В карте {} ЗЕТ.'.format(sum_normal, sum_zet)
            print(errors)
            return_err_arr.append(errors)
    # ### ------------------------------------ ###
    return return_err_arr