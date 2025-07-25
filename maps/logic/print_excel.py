import io
import os
from math import floor

import openpyxl
import xlsxwriter
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)

from maps.logic.take_from_bd import create_json_print, elective_disciplines, get_default_shortcuts, get_user_shortcuts
from maps.logic.tools import get_maximum_rows
from maps.models import (AupInfo, AupData, Groups, D_Period)

ROW_START_DISCIPLINES = 4
ROW_HEIGHT = 23
COLUMN_WIDTH = 46
SUM_ROW_HEIGHT = ROW_HEIGHT * 30
SUM_COLUMN_WIDTH = COLUMN_WIDTH * 8

border_thin = Side(style='thin', color='000000')
border_thick = Side(style='thick', color='000000')


def makeLegend(wb, table, aup):
    ws = wb.create_sheet('Legend')

    ws['A1'].value = 'ЗЕТ'
    ws['A1'].style = 'standart'
    ws['B1'].value = 'Группа'
    ws['B1'].style = 'standart'
    ws.column_dimensions["A"].width = 25.0
    ws.column_dimensions["B"].width = 60.0

    # Словарь с ключом id группировки и значением - сумма зет в карте для группировки
    table_dict = {}

    # Словарь для хранения всех группировок
    group_dict = {}

    for group in Groups.query.all():
        group_dict[group.id_group] = {'name': group.name_group, 'color': group.color}

    # Считаем сумму зет для каждой группировки
    for column in table:
        for item in column:
            if item['id_group'] in table_dict:
                table_dict[item['id_group']] += item['zet']
            else:
                table_dict[item['id_group']] = item['zet']
    sum_zet = 0

    for i, key_value in enumerate(table_dict.items()):
        ws.row_dimensions[i + 2].height = 20
        ws['A' + str(i + 2)].value = int(key_value[1])
        ws['A' + str(i + 2)].style = 'standart'
        ws['B' + str(i + 2)].value = group_dict[key_value[0]]['name']
        sum_zet += int(key_value[1])

        color_text_cell(ws, 'B' + str(i + 2), group_dict[key_value[0]]['color'].replace('#', ''))

    ws['A' + str(len(table_dict) + 2)].style = 'standart'
    ws['A' + str(len(table_dict) + 2)].value = 'Итого: ' + str(sum_zet)

    last_row = len(table_dict) + 2
    # отступ от Итого
    last_row += 4
    # --- часть легенды с факультативами ---

    ws['B' + str(last_row)].value = 'Факультативы:'
    ws['B' + str(last_row)].style = 'standart'
    last_row += 1
    ws['A' + str(last_row)].style = 'standart'
    ws['B' + str(last_row)].style = 'standart'
    ws['A' + str(last_row)].value = 'ЗЕТ'
    ws['B' + str(last_row)].value = 'Название'
    ws['C' + str(last_row)].value = 'Часы'
    ws['C' + str(last_row)].style = 'standart'

    dis = elective_disciplines(aup)
    last_row += 1
    sum = 0
    # запись названий в столбец B
    for key, value in dis.items():
        ws['B' + str(last_row)] = key
        ws['B' + str(last_row)].style = 'standart'
        ws['C' + str(last_row)] = value
        ws['C' + str(last_row)].style = 'standart'
        ws['A' + str(last_row)] = value / 36
        ws['A' + str(last_row)].style = 'standart'
        sum += value / 36
        last_row += 1

    ws['A' + str(last_row)].style = 'standart'
    ws['A' + str(last_row)].value = f'Итого: {sum}'


def saveMap(aup, static, papper_size, orientation, control: bool = False, load: bool = False):
    aup = AupInfo.query.filter_by(num_aup=aup).first()
    data = AupData.query.filter_by(id_aup=aup.id_aup).order_by(AupData.shifr, AupData.id_discipline,
                                                               AupData.id_period).all()
    filename_map = aup.file
    filename_map_down = f"КД {filename_map}"
    filename_map = os.path.join(static, 'temp', f"КД {filename_map}")

    table = create_json_print(data)

    max_zet = find_max_zet_excel(table)
    table = add_table_to_arr_and_sort(table['data'])
    ws, wb = CreateMap(filename_map, max_zet, len(table))

    header = Header(aup)
    header1 = f'''КАРТА ДИСЦИПЛИН УЧЕБНОГО ПЛАНА
{header[0]}  
Профиль "{header[1]}", {header[2]} год набора, {header[3]}'''
    ws['A1'].style = 'header'
    ws['A1'] = header1

    ws.merge_cells(
        f'A1:{chr(ord("A") + len(table))}1')

    for width_border in range(1, len(table) + 1):
        ws[f"{chr(ord('A') + width_border)}1"].style = 'special'
        ws[f"{chr(ord('A') + width_border)}2"].style = 'special'

    for course in range(floor(len(table) / 2)):
        ws[chr(ord("B") + course * 2) + f"{ROW_START_DISCIPLINES - 2}"] = str(course + 1) + " курс"
        ws[chr(ord("B") + course * 2) + f"{ROW_START_DISCIPLINES - 2}"].style = 'special'
        ws.merge_cells(
            f'{chr(ord("B") + course * 2)}{ROW_START_DISCIPLINES - 2}:{chr(ord("B") + course * 2 + 1)}{ROW_START_DISCIPLINES - 2}')

    if not (len(table) / 2).is_integer():
        ws[chr(ord("B") + floor(len(table) / 2) * 2) + f"{ROW_START_DISCIPLINES - 2}"] = str(
            floor(len(table) / 2) + 1) + " курс"
        ws[chr(ord("B") + floor(len(table) / 2) * 2) + f"{ROW_START_DISCIPLINES - 2}"].style = 'special'

    for semester in range(len(table)):
        ws[chr(ord("B") + semester) + f"{ROW_START_DISCIPLINES - 1}"] = str(semester + 1) + ' семестр'
        ws[chr(ord("B") + semester) + f"{ROW_START_DISCIPLINES - 1}"].style = 'special'

    for i in range(len(table)):
        merged = 0
        for el in table[i]:
            column = chr(ord("B") + i)
            cell = f"{column}{ROW_START_DISCIPLINES + merged}"

            ws[cell] = el['discipline']
            
            if any([load, control]):
                ws[cell] += load_and_control(el, load, control)
              
            color = el['color'].replace('#', '')

            color_text_cell(ws, cell, color)

            if el['zet'] < 1:
                el['zet'] = 1.0

            merge_range = f"{cell}:{column}{ROW_START_DISCIPLINES + merged + round(el['zet'] * 2) - 1}"
            ws.merge_cells(merge_range)

            merged += round(el['zet'] * 2)

    makeLegend(wb, table, aup)

    set_print_properties(table, ws, max_zet)

    ### Установить нижний колонтитул
    ws.oddFooter.right.text = f"АУП {aup.num_aup}"
    ws.oddFooter.right.size = 14
    ws.oddFooter.right.font = "Arial,Bold"
    ws.oddFooter.right.color = "000000"
    if papper_size == "3":
        ws.page_setup.papperSize = ws.PAPERSIZE_A3
    elif papper_size == "4":
        ws.page_setup.papperSize = ws.PAPERSIZE_A4
    if orientation == "land":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    elif orientation == "port":
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    for sheets in wb:
        sheets.sheet_view.zoomScale = 45

    wb.save(filename=filename_map)
    return filename_map


def color_text_cell(ws, cell, color):
    ws[cell].style = 'standart'
    r, g, b = tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))
    gray = (r + g + b) / 3

    if gray < 140:
        ws[cell].font = Font(bold=True, size=16, name='Arial', color="FFFFFF")  # Dvorf False 18
    else:
        ws[cell].font = Font(bold=True, size=16, name='Arial', color="000000")  # Dvorf False 18

    ws[cell].fill = PatternFill(start_color=str(
        color), end_color=str(color), fill_type='solid')


def find_max_zet_excel(table):
    max_zet = 0
    terms = {}
    for item in table['data']:
        if item['num_col'] not in terms:
            terms[item['num_col']] = item['zet']
            continue
        terms[item['num_col']] += item['zet']
    for _, value in terms.items():
        if value > max_zet:
            max_zet = value
    return int(max_zet)


def set_print_properties(table, ws, max_zet):
    # Set properties
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    max_row = get_maximum_rows(sheet_object=ws) * 2
    ws.print_area = 'A1:' + str(alphabet[len(table)]) + str(max_zet * 2 + ROW_START_DISCIPLINES - 1)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=1 / 5, right=1 / 5, top=1 / 5, bottom=1 / 5, header=0.1, footer=0.1)  # Правка на 0,5 см

    # Высота строк где дисциплины
    for height_row in range(ROW_START_DISCIPLINES, max_zet * 2 + ROW_START_DISCIPLINES):
        ws.row_dimensions[height_row].height = SUM_ROW_HEIGHT / max_zet

    # Ширины столбцов где дисциплины
    for width_column in range(1, len(table) + 1):
        ws.column_dimensions[f'{chr(ord("A") + width_column)}'].width = SUM_COLUMN_WIDTH / len(table)

    # Определяем высоту наибольшего столбца(семестра) для избежания выхода за границы
    col_max_height = max([sum(map(lambda cell: cell['zet'], col)) * 2 for col in table])

    # Перевод численных координат в координаты с буквой
    cell = lambda x, y: f'{chr(ord("A") + x + 1)}{ROW_START_DISCIPLINES + y}'

    # Пустая ли ячейка
    is_none = lambda x, y: cell(x, y) not in ws.merged_cells

    for x in range(len(table)):
        for y in range(int(col_max_height)):
            border = Border()
            if (not is_none(x, y)):
                border = Border(left=border_thin, right=border_thin, bottom=border_thin)

            if (x + 1 == len(table)):
                border.right = border_thick

            if (y + 1 == col_max_height):
                border.bottom = border_thick

            ws[cell(x, y)].border = border

    ws.column_dimensions['A'].width = 10  # Column ZET
    ### Установить открытие страницы полностью
    ws.page_setup.fitToPage = True


def Header(aup):
    """
        Возвращает данные для шапки карты
    """
    year_begin = aup.year_beg
    program = aup.spec.program_code + ' ' + aup.spec.okco.name_okco
    form = aup.form.form + " форма обучения"
    spec = aup.spec.name_spec
    # date_file = aup.file.split(' ')[-4]
    return [program, spec, year_begin, form]


def addStyles(workbook):
    """
        Создает и добавляет стили к файлу Excel
    """

    ns_standart = NamedStyle(name='standart')
    ns_standart.font = Font(bold=True, size=16, name='Arial')
    ns_standart.border = Border(left=border_thin, top=border_thin, right=border_thin, bottom=border_thin)
    ns_standart.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns_standart)

    ns_special = NamedStyle(name='special')
    ns_special.font = Font(bold=True, size=16, name='Arial')
    ns_special.border = Border(left=border_thick, top=border_thick, right=border_thick, bottom=border_thick)
    ns_special.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns_special)

    ns_header = NamedStyle(name='header')
    ns_header.font = Font(bold=True, size=22, name='Arial')
    ns_header.border = Border(left=border_thick, top=border_thick, right=border_thick, bottom=border_thick)
    ns_header.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns_header)

    standart_last_right = NamedStyle(name='standart_last_right')
    standart_last_right.font = Font(bold=True, size=16, name='Arial')
    standart_last_right.border = Border(left=border_thin, top=border_thin, right=border_thick, bottom=border_thin)
    standart_last_right.alignment = Alignment(
        horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(standart_last_right)


def CreateMap(filename_map, max_zet, table_length):
    """
        Функция создает карту и задает все данные кроме предметов в семестрах, на вход требует имя карты
    """

    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 40, 40)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active

    addStyles(workbook)

    worksheet.row_dimensions[1].height = 90
    worksheet.row_dimensions[2].height = 20

    merge_range = f"A{ROW_START_DISCIPLINES - 2}:A{ROW_START_DISCIPLINES - 1}"
    worksheet.merge_cells(merge_range)
    worksheet[f"A{ROW_START_DISCIPLINES - 2}"].style = 'special'
    worksheet[f"A{ROW_START_DISCIPLINES - 1}"].style = 'special'
    worksheet[f"A{ROW_START_DISCIPLINES - 2}"] = 'З.Е.'

    for col in range(1, max_zet + 1):
        merge_range = f"A{col * 2 + ROW_START_DISCIPLINES - 2}:A{col * 2 + ROW_START_DISCIPLINES - 1}"
        worksheet.merge_cells(merge_range)
        worksheet["A" + str(col * 2 + ROW_START_DISCIPLINES - 2)] = col
        worksheet["A" + str(col * 2 + ROW_START_DISCIPLINES - 2)].style = 'special'
        worksheet["A" + str(col * 2 + ROW_START_DISCIPLINES - 1)].style = 'special'
    for col in range(ord('B'), ord('C')):
        worksheet[chr(col) + str(5)].style = 'special'
    return worksheet, workbook


def add_table_to_arr_and_sort(table):
    count_columns = 0
    for item in table:
        if item['num_col'] > count_columns:
            count_columns = item['num_col']

    new_table = [[] for _ in range(count_columns)]

    for item in table:
        new_table[item['num_col'] - 1].append(item)

    for a in range(len(new_table)):
        for i in range(len(new_table[a]) - 1):
            for j in range(len(new_table[a]) - i - 1):
                if new_table[a][j]['num_row'] > new_table[a][j + 1]['num_row']:
                    new_table[a][j], new_table[a][j + 1] = new_table[a][j + 1], new_table[a][j]
    return (new_table)


def get_aup_data_excel(aup: str) -> tuple[io.BytesIO, str]:
    in_memory_file = io.BytesIO()

    book = xlsxwriter.Workbook(in_memory_file)
    sheet = book.add_worksheet("Лист1")

    header_format = book.add_format()
    header_format.set_bold(True)

    sheet.write_row(0, 0, ['Наименование', "Содержание", ], header_format)
    sheet.write_column(1, 0, [
        'Номер АУП',
        'Вид образования',
        'Уровень образования',
        'Направление (специальность)',
        'Код специальности',
        'Квалификация',
        'Профиль (специализация)',
        'Тип стандарта',
        'Факультет',
        'Выпускающая кафедра',
        'Форма обучения',
        'Год набора',
        'Период обучения',
        'На базе',
        'Фактический срок обучения',
    ])

    aup_info: AupInfo = AupInfo.query.filter_by(num_aup=aup).first()

    sheet.write_column(1, 1, [
        aup_info.num_aup,
        aup_info.type_educ,
        aup_info.degree.name_deg,
        aup_info.spec.okco.name_okco,
        aup_info.spec.program_code,
        aup_info.qualification,
        aup_info.spec.name_spec,
        aup_info.type_standard,
        aup_info.faculty.name_faculty,
        aup_info.department.name_department,
        aup_info.form.form,
        str(aup_info.year_beg),
        aup_info.period_educ,
        aup_info.base,
        F"{aup_info.years} года" + (f' {aup_info.months} месяцев' if aup_info.months else ''),
    ])

    sheet.autofit()

    sheet = book.add_worksheet("Лист2")

    sheet.write_row(0, 0, [
        'Блок',
        'Шифр',
        'Часть',
        'Модуль',
        'Тип записи',
        'Дисциплина',
        'Период контроля',
        'Нагрузка',
        'Количество',
        'Ед. изм.',
        'ЗЕТ',
    ], header_format)

    periods = {el.id: el.title for el in D_Period.query.all()}

    for i, el in enumerate(aup_info.aup_data, start=1):
        el: AupData
        sheet.write_row(i, 0, [
            el.block.title,
            el.shifr,
            el.part.title,
            el.module.title,
            el.type_record.title,
            el.discipline.title,
            periods[el.id_period],
            el.type_control.title,
            el.amount / 100,
            el.ed_izmereniya.title,
            el.zet / 100,
        ])

    sheet.autofit()
    book.close()
    return in_memory_file, F"{aup_info.num_aup} {aup_info.degree.name_deg} {aup_info.spec.name_spec} {aup_info.form.form}"


def load_and_control(el, load: bool, control: bool):
    '''
    Форматирование нагрузки и контроля
    '''
    cuts = get_default_shortcuts()
    control_result = "\n"
    value = "\n\n"
    if control:
        for element in el['type']['session']:
            temp = int(element['control_type_id'])
            control_result += (temp if temp not in cuts else cuts[temp])
    if load:        
        temp = ""
        for element in el['type']['value']:
            temp = int(element['control_type_id'])
            if (temp == 7): 
                control_result += (", КП " if control else " КП ")
            else:    
                value += (temp if temp not in cuts else cuts[temp]) + " " + str(int(element['amount'])) + " " + str(element['amount_type']) + "\t"

    return value + control_result
